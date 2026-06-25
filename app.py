"""
LuxeLocks Hub - 轻量订单管理中台
Shopify + MCP + 海外仓
"""

import os
import json
import hashlib
import hmac
import base64
import sqlite3
import asyncio
import smtplib
import threading
import time
import re
import zipfile
from io import BytesIO
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from contextlib import asynccontextmanager
from pathlib import Path

import httpx
from fastapi import FastAPI, Request, HTTPException, Query
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi import UploadFile, File
import shutil
from fastapi.templating import Jinja2Templates

# ────────────────────────────────────────
# 配置
# ────────────────────────────────────────
BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / "luxelocks.db"

from dotenv import load_dotenv
load_dotenv(BASE_DIR / ".env")

SHOPIFY_SHOP = os.getenv("SHOPIFY_SHOP", "your-store.myshopify.com")
SHOPIFY_API_KEY = os.getenv("SHOPIFY_API_KEY", "")
SHOPIFY_API_PASSWORD = os.getenv("SHOPIFY_API_PASSWORD", "")
SHOPIFY_ACCESS_TOKEN = os.getenv("SHOPIFY_ACCESS_TOKEN", "")
SHOPIFY_WEBHOOK_SECRET = os.getenv("SHOPIFY_WEBHOOK_SECRET", "")

# 海外仓 API 配置 (领星 OMS)
OMS_APP_KEY = os.getenv("OMS_APP_KEY", "94611d271e3546c5a0e513f981d256df")
OMS_APP_SECRET = os.getenv("OMS_APP_SECRET", "8774e4155d3f47da899cafe77e8bd0c5")
OMS_DOMAIN = os.getenv("OMS_DOMAIN", "oms.xlwms.com")

# 产品媒体文件管理
MEDIA_ROOT = Path(os.getenv("MEDIA_ROOT", r"\\huawei\Users\HUAWEI\Pictures\产品"))
# 媒体文件树缓存（避免每次请求扫描UNC路径 ~7s）
_media_tree_cache = {"paths": [], "ts": 0}
_MEDIA_CACHE_TTL = 300  # 5分钟过期，点同步按钮强制刷新
# 图片代理缓存目录
_IMG_CACHE_DIR = BASE_DIR / ".img_cache"
_IMG_CACHE_DIR.mkdir(exist_ok=True)
# Shopify location ID 缓存
_shopify_location_id = None

def _get_media_root():
    """获取媒体根路径（懒解析，避免UNC不可用时import崩溃）"""
    try:
        return MEDIA_ROOT.resolve()
    except Exception:
        return MEDIA_ROOT

# 邮件提醒配置
EMAIL_HOST = "smtp.qq.com"
EMAIL_PORT = 587
EMAIL_USER = os.getenv("EMAIL_USER", "gkway@qq.com")
EMAIL_PASS = os.getenv("EMAIL_PASS", "warlyhedcwjfbaib")
EMAIL_TO = os.getenv("EMAIL_TO", "gkway@qq.com")


# ────────────────────────────────────────
# 数据库初始化
# ────────────────────────────────────────
def init_db():
    conn = sqlite3.connect(str(DB_PATH), timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            platform TEXT NOT NULL DEFAULT 'shopify',
            platform_order_id TEXT NOT NULL,
            order_number TEXT,
            customer_name TEXT,
            customer_email TEXT,
            total_price REAL,
            currency TEXT DEFAULT 'USD',
            status TEXT DEFAULT 'pending',
            shipping_address TEXT,
            tracking_number TEXT,
            tracking_company TEXT,
            line_items TEXT,
            raw_data TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(platform, platform_order_id)
        );

        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            platform TEXT NOT NULL DEFAULT 'shopify',
            platform_product_id TEXT NOT NULL,
            sku TEXT,
            title TEXT,
            variant_title TEXT,
            price REAL,
            inventory_quantity INTEGER DEFAULT 0,
            image_url TEXT,
            raw_data TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(platform, platform_product_id, sku)
        );

        CREATE TABLE IF NOT EXISTS inventory_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sku TEXT NOT NULL,
            quantity_change INTEGER,
            reason TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS sync_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            platform TEXT,
            action TEXT,
            status TEXT,
            detail TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE INDEX IF NOT EXISTS idx_orders_platform ON orders(platform, platform_order_id);
        CREATE INDEX IF NOT EXISTS idx_orders_status ON orders(status);
        CREATE INDEX IF NOT EXISTS idx_orders_created ON orders(created_at);
        CREATE INDEX IF NOT EXISTS idx_products_sku ON products(sku);
    """)
    try: conn.execute("ALTER TABLE products ADD COLUMN oms_available_qty INTEGER DEFAULT 0")
    except: pass
    try: conn.execute("ALTER TABLE products ADD COLUMN oms_transit_qty INTEGER DEFAULT 0")
    except: pass
    try: conn.execute("ALTER TABLE products ADD COLUMN product_status TEXT DEFAULT 'active'")
    except: pass
    try: conn.execute("ALTER TABLE products ADD COLUMN variant_id TEXT DEFAULT ''")
    except: pass
    try: conn.execute("ALTER TABLE products ADD COLUMN inventory_item_id TEXT DEFAULT ''")
    except: pass
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS headhaul_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_order_no TEXT,
            transfer_order_no TEXT,
            tracking_number TEXT,
            shipping_channel TEXT,
            status TEXT DEFAULT 'pending',
            origin TEXT,
            destination TEXT,
            weight REAL,
            pieces INTEGER DEFAULT 1,
            raw_data TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_headhaul_customer ON headhaul_orders(customer_order_no);
        CREATE INDEX IF NOT EXISTS idx_headhaul_transfer ON headhaul_orders(transfer_order_no);

        CREATE TABLE IF NOT EXISTS important_matters (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            content TEXT,
            remind_at TEXT NOT NULL,
            status TEXT DEFAULT 'pending',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS warehouse_inventory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            goods_id TEXT,
            goods_name TEXT,
            reference_code TEXT,
            status TEXT,
            warehouse_name TEXT NOT NULL,
            total_inventory INTEGER DEFAULT 0,
            available_inventory INTEGER DEFAULT 0,
            defective INTEGER DEFAULT 0,
            good_reserved INTEGER DEFAULT 0,
            in_transit_total INTEGER DEFAULT 0,
            in_transit_shipped INTEGER DEFAULT 0,
            in_transit_receiving INTEGER DEFAULT 0,
            in_transit_returning INTEGER DEFAULT 0,
            platform TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_warehouse_name ON warehouse_inventory(warehouse_name);
        CREATE INDEX IF NOT EXISTS idx_warehouse_ref ON warehouse_inventory(reference_code);
    """)
    # SKU映射表: 解决不同系统简化SKU不一致问题
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS sku_mapping (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shopify_simple_sku TEXT NOT NULL UNIQUE,
            cross_border_simple_sku TEXT,
            luxelocks_simple_sku TEXT,
            velourahair_simple_sku TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS reminders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            content TEXT,
            remind_date TEXT NOT NULL,
            email TEXT,
            sent INTEGER DEFAULT 0,
            repeat_type TEXT DEFAULT '',
            repeat_day INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    try: conn.execute("ALTER TABLE reminders ADD COLUMN repeat_type TEXT DEFAULT ''")
    except: pass
    try: conn.execute("ALTER TABLE reminders ADD COLUMN repeat_day INTEGER DEFAULT 0")
    except: pass
    # TikTok SKU映射
    conn.execute("""
        CREATE TABLE IF NOT EXISTS tiktok_sku_mapping (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tiktok_product_id TEXT NOT NULL UNIQUE,
            sku TEXT NOT NULL,
            simple_sku TEXT,
            price REAL DEFAULT 0,
            product_cost REAL DEFAULT 0,
            shipping_fee REAL DEFAULT 0,
            platform_fee_rate REAL DEFAULT 0.06,
            ad_cost REAL DEFAULT 0,
            refund_loss REAL DEFAULT 0,
            return_rate REAL DEFAULT 0.20,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    for sql in [
        "ALTER TABLE tiktok_sku_mapping ADD COLUMN simple_sku TEXT",
        "ALTER TABLE tiktok_sku_mapping ADD COLUMN price REAL DEFAULT 0",
        "ALTER TABLE tiktok_sku_mapping ADD COLUMN product_cost REAL DEFAULT 0",
        "ALTER TABLE tiktok_sku_mapping ADD COLUMN shipping_fee REAL DEFAULT 0",
        "ALTER TABLE tiktok_sku_mapping ADD COLUMN platform_fee_rate REAL DEFAULT 0.06",
        "ALTER TABLE tiktok_sku_mapping ADD COLUMN ad_cost REAL DEFAULT 0",
        "ALTER TABLE tiktok_sku_mapping ADD COLUMN refund_loss REAL DEFAULT 0",
        "ALTER TABLE tiktok_sku_mapping ADD COLUMN return_rate REAL DEFAULT 0.20",
        "ALTER TABLE tiktok_sku_mapping ADD COLUMN notes TEXT",
    ]:
        try: conn.execute(sql)
        except: pass
    # 采购资源信息池
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS procurement_resources (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            simple_sku TEXT NOT NULL UNIQUE,
            image_url TEXT,
            supplier TEXT,
            cost REAL,
            estimated_delivery TEXT,
            purchase_link TEXT,
            contact_name TEXT,
            wechat TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    try:
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_procurement_sku_uq ON procurement_resources(simple_sku)")
    except:
        pass
    # 供应商管理
    conn.execute("""
        CREATE TABLE IF NOT EXISTS suppliers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            seller_id TEXT,
            purchase_link TEXT,
            contact_name TEXT,
            wechat TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    try: conn.execute("ALTER TABLE suppliers ADD COLUMN purchase_link TEXT DEFAULT ''")
    except: pass
    try: conn.execute("ALTER TABLE suppliers ADD COLUMN contact_name TEXT DEFAULT ''")
    except: pass
    try: conn.execute("ALTER TABLE suppliers ADD COLUMN wechat TEXT DEFAULT ''")
    except: pass
    try: conn.execute("ALTER TABLE procurement_resources ADD COLUMN seller_id TEXT DEFAULT ''")
    except: pass
    # TikTok Wig Ops 中控模块
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS tiktok_skus (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sku TEXT NOT NULL UNIQUE,
            product_name TEXT,
            color TEXT,
            length TEXT,
            price REAL DEFAULT 23.99,
            product_cost REAL DEFAULT 0,
            shipping_fee REAL DEFAULT 0,
            platform_fee_rate REAL DEFAULT 0.06,
            ad_cost REAL DEFAULT 0,
            return_rate REAL DEFAULT 0.20,
            refund_loss REAL DEFAULT 0,
            stock INTEGER DEFAULT 0,
            daily_sales REAL DEFAULT 0,
            lead_time_days INTEGER DEFAULT 30,
            safety_stock INTEGER DEFAULT 10,
            status TEXT DEFAULT 'testing',
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS tiktok_videos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            account_name TEXT,
            sku TEXT,
            tiktok_video_id TEXT,
            tiktok_product_id TEXT,
            product_name TEXT,
            creator_id TEXT,
            publish_date TEXT,
            video_angle TEXT,
            hook TEXT,
            selling_points TEXT,
            display_order TEXT,
            voiceover TEXT,
            cover_text TEXT,
            caption TEXT,
            hashtags TEXT,
            posted INTEGER DEFAULT 0,
            views INTEGER DEFAULT 0,
            likes_count INTEGER DEFAULT 0,
            comments_count INTEGER DEFAULT 0,
            shares_count INTEGER DEFAULT 0,
            product_impressions INTEGER DEFAULT 0,
            product_clicks INTEGER DEFAULT 0,
            orders INTEGER DEFAULT 0,
            gmv REAL DEFAULT 0,
            video_ctr REAL DEFAULT 0,
            completion_rate REAL DEFAULT 0,
            gpm REAL DEFAULT 0,
            comments TEXT,
            platform_diagnosis TEXT,
            diagnosis TEXT,
            repeat_action TEXT,
            source_file TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE INDEX IF NOT EXISTS idx_tiktok_skus_sku ON tiktok_skus(sku);
        CREATE INDEX IF NOT EXISTS idx_tiktok_videos_sku ON tiktok_videos(sku);
        CREATE INDEX IF NOT EXISTS idx_tiktok_videos_date ON tiktok_videos(publish_date);

        CREATE TABLE IF NOT EXISTS tiktok_video_performance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_key TEXT NOT NULL UNIQUE,
            creator_nickname TEXT,
            creator_id TEXT,
            video_info TEXT,
            video_id TEXT,
            publish_time TEXT,
            product_name TEXT,
            tiktok_product_id TEXT,
            sku TEXT,
            simple_sku TEXT,
            vv INTEGER DEFAULT 0,
            likes_count INTEGER DEFAULT 0,
            comments_count INTEGER DEFAULT 0,
            shares_count INTEGER DEFAULT 0,
            new_followers_count INTEGER DEFAULT 0,
            product_redirects INTEGER DEFAULT 0,
            product_impressions INTEGER DEFAULT 0,
            product_clicks INTEGER DEFAULT 0,
            unique_customers INTEGER DEFAULT 0,
            attributed_sku_orders INTEGER DEFAULT 0,
            video_sku_orders INTEGER DEFAULT 0,
            indirect_sku_orders INTEGER DEFAULT 0,
            attributed_units INTEGER DEFAULT 0,
            product_units INTEGER DEFAULT 0,
            indirect_units INTEGER DEFAULT 0,
            attributed_gmv REAL DEFAULT 0,
            video_gmv REAL DEFAULT 0,
            indirect_gmv REAL DEFAULT 0,
            gpm REAL DEFAULT 0,
            video_ctr REAL DEFAULT 0,
            redirect_rate REAL DEFAULT 0,
            completion_rate REAL DEFAULT 0,
            ctor_sku_orders REAL DEFAULT 0,
            platform_diagnosis TEXT,
            local_diagnosis TEXT,
            repeat_action TEXT,
            source_file TEXT,
            imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_tiktok_video_perf_video ON tiktok_video_performance(video_id);
        CREATE INDEX IF NOT EXISTS idx_tiktok_video_perf_product ON tiktok_video_performance(tiktok_product_id);
        CREATE INDEX IF NOT EXISTS idx_tiktok_video_perf_sku ON tiktok_video_performance(sku);
    """)
    for sql in [
        "ALTER TABLE tiktok_videos ADD COLUMN tiktok_video_id TEXT",
        "ALTER TABLE tiktok_videos ADD COLUMN tiktok_product_id TEXT",
        "ALTER TABLE tiktok_videos ADD COLUMN product_name TEXT",
        "ALTER TABLE tiktok_videos ADD COLUMN creator_id TEXT",
        "ALTER TABLE tiktok_videos ADD COLUMN likes_count INTEGER DEFAULT 0",
        "ALTER TABLE tiktok_videos ADD COLUMN comments_count INTEGER DEFAULT 0",
        "ALTER TABLE tiktok_videos ADD COLUMN shares_count INTEGER DEFAULT 0",
        "ALTER TABLE tiktok_videos ADD COLUMN product_impressions INTEGER DEFAULT 0",
        "ALTER TABLE tiktok_videos ADD COLUMN video_ctr REAL DEFAULT 0",
        "ALTER TABLE tiktok_videos ADD COLUMN completion_rate REAL DEFAULT 0",
        "ALTER TABLE tiktok_videos ADD COLUMN gpm REAL DEFAULT 0",
        "ALTER TABLE tiktok_videos ADD COLUMN platform_diagnosis TEXT",
        "ALTER TABLE tiktok_videos ADD COLUMN source_file TEXT",
    ]:
        try: conn.execute(sql)
        except: pass
    try:
        conn.execute("CREATE INDEX IF NOT EXISTS idx_tiktok_videos_tiktok_id ON tiktok_videos(tiktok_video_id, tiktok_product_id)")
    except:
        pass
    # ── 迁移: 为 orders 添加 shipping_type 字段 ──
    try:
        conn.execute("ALTER TABLE orders ADD COLUMN shipping_type TEXT DEFAULT '国内直发'")
    except:
        pass  # 字段已存在

    # 为已存在的订单计算运输类型（跳过已发货的，避免覆盖历史快照）
    existing = conn.execute(
        "SELECT id, line_items, status, shipping_type FROM orders"
        " WHERE (shipping_type IS NULL OR shipping_type = '' OR shipping_type = '国内直发')"
        " AND status != 'shipped'"
    ).fetchall()
    if existing:
        # 加载跨境库存
        wh_items = conn.execute(
            "SELECT reference_code, available_inventory FROM warehouse_inventory"
            " WHERE warehouse_name IN ('CrossBorder', 'NewProducts')"
        ).fetchall()
        cross_stock = {}
        for w in wh_items:
            ref = simplify_sku(w["reference_code"])
            if ref:
                cross_stock[ref] = cross_stock.get(ref, 0) + w["available_inventory"]

        mappings = conn.execute(
            "SELECT shopify_simple_sku, cross_border_simple_sku FROM sku_mapping"
        ).fetchall()
        map_cross = {}
        for m in mappings:
            if m["cross_border_simple_sku"]:
                map_cross[m["shopify_simple_sku"]] = m["cross_border_simple_sku"]

        for row in existing:
            items = json.loads(row["line_items"] or "[]")
            stype = "国内直发"
            for item in items:
                ss = simplify_sku(item.get("sku", ""))
                cross_key = map_cross.get(ss, ss)
                if cross_stock.get(cross_key, 0) > 0:
                    stype = "海外仓发货"
                    break
            conn.execute("UPDATE orders SET shipping_type = ? WHERE id = ?", (stype, row["id"]))
    conn.commit()
    conn.close()


def get_db():
    conn = sqlite3.connect(str(DB_PATH), timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout=30000")
    return conn


def safe_float(value, default=0.0):
    try:
        if value in (None, ""):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def safe_int(value, default=0):
    try:
        if value in (None, ""):
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


TIKTOK_DEFAULT_ACCOUNTS = ["official", "channel01", "channel02", "channel03", "channel04"]
TIKTOK_ANGLE_POOL = [
    "整体效果", "发顶自然", "颜色自然光", "侧面层次", "背面发量",
    "可调节内网", "新手友好", "23.99包邮"
]


def calc_tiktok_profit(row):
    price = safe_float(row.get("price"), 23.99)
    product_cost = safe_float(row.get("product_cost"))
    shipping_fee = safe_float(row.get("shipping_fee"))
    platform_fee_rate = safe_float(row.get("platform_fee_rate"), 0.06)
    ad_cost = safe_float(row.get("ad_cost"))
    return_rate = safe_float(row.get("return_rate"), 0.20)
    refund_loss = safe_float(row.get("refund_loss")) or (product_cost + shipping_fee)
    platform_fee = price * platform_fee_rate
    base_profit = price - product_cost - shipping_fee - platform_fee - ad_cost
    net_profit = base_profit - refund_loss * return_rate
    profit_rate = net_profit / price if price else 0
    denominator = max(1 - platform_fee_rate, 0.01)
    break_even_price = (product_cost + shipping_fee + ad_cost + refund_loss * return_rate) / denominator
    suggested_price = max(23.99, break_even_price + 6)
    stock = safe_int(row.get("stock"))
    daily_sales = safe_float(row.get("daily_sales"))
    lead_time_days = safe_int(row.get("lead_time_days"), 30)
    safety_stock = safe_int(row.get("safety_stock"), 10)
    days_left = None
    reorder_in_days = None
    stock_status = "观察"
    if daily_sales > 0:
        days_left = stock / daily_sales
        reorder_in_days = days_left - lead_time_days
        if days_left <= lead_time_days or stock <= safety_stock:
            stock_status = "需补货"
        elif days_left <= lead_time_days + 7:
            stock_status = "预警"
        else:
            stock_status = "安全"
    elif stock <= safety_stock:
        stock_status = "低库存"

    if net_profit < 3:
        profit_label = "高风险"
    elif net_profit >= 6:
        profit_label = "可主推"
    else:
        profit_label = "可测试"

    return {
        "platform_fee": round(platform_fee, 2),
        "base_profit": round(base_profit, 2),
        "net_profit": round(net_profit, 2),
        "profit_rate": round(profit_rate * 100, 1),
        "break_even_price": round(break_even_price, 2),
        "suggested_price": round(suggested_price, 2),
        "profit_label": profit_label,
        "days_left": round(days_left, 1) if days_left is not None else None,
        "reorder_in_days": round(reorder_in_days, 1) if reorder_in_days is not None else None,
        "stock_status": stock_status
    }


def diagnose_tiktok_video(row):
    views = safe_int(row.get("views"))
    clicks = safe_int(row.get("product_clicks"))
    orders = safe_int(row.get("orders"))
    gmv = safe_float(row.get("gmv"))
    ctr = clicks / views if views else 0
    conversion = orders / clicks if clicks else 0
    comments = (row.get("comments") or "").lower()

    diagnosis = "数据不足"
    repeat_action = "继续记录"
    if views >= 1000 and ctr < 0.01:
        diagnosis = "播放高点击低"
        repeat_action = "重做封面和前3秒卖点"
    elif clicks >= 20 and conversion < 0.03:
        diagnosis = "点击高成交低"
        repeat_action = "检查价格、详情页、评价和评论信任"
    elif views < 500 and orders > 0:
        diagnosis = "小流量高转化"
        repeat_action = "换账号复拍同结构"
    elif views >= 1000 and orders > 0:
        diagnosis = "可复拍模板"
        repeat_action = "保留开头和展示顺序连续复拍"
    elif views >= 500 and clicks >= 10 and orders == 0:
        diagnosis = "有兴趣未成交"
        repeat_action = "增加上头效果、发顶近景和包邮价格"
    if "color" in comments or "颜色" in comments:
        repeat_action = "补拍自然光颜色对比"
    if "wear" in comments or "beginner" in comments or "新手" in comments:
        repeat_action = "补拍新手佩戴步骤"

    return {
        "ctr": round(ctr * 100, 2),
        "conversion": round(conversion * 100, 2),
        "aov": round(gmv / orders, 2) if orders else 0,
        "diagnosis": diagnosis,
        "repeat_action": repeat_action
    }


def build_tiktok_script(sku_info, angle):
    sku = sku_info.get("sku") or "SKU"
    color = sku_info.get("color") or "natural color"
    length = sku_info.get("length") or ""
    price = safe_float(sku_info.get("price"), 23.99)
    product_name = sku_info.get("product_name") or f"{length} {color} wig".strip()
    angle_map = {
        "整体效果": ("full look", "full look -> face frame -> price", "This whole look is ready in seconds."),
        "发顶自然": ("natural top", "top close-up -> hand parting -> full look", "Look how natural the top looks up close."),
        "颜色自然光": ("color in natural light", "window light -> side move -> full look", "This color looks even better in natural light."),
        "侧面层次": ("side layers", "side profile -> hair movement -> full look", "The side layers make it look soft and real."),
        "背面发量": ("back view", "back view -> turn around -> full look", "The back has enough volume without looking heavy."),
        "可调节内网": ("adjustable cap", "inside cap -> straps -> finished look", "The adjustable cap makes it beginner friendly."),
        "新手友好": ("beginner friendly", "before -> put on -> final look", "No install skills needed for this everyday wig."),
        "23.99包邮": ("$23.99 free shipping", "full look -> close-up -> price CTA", "A cute everyday wig for only $23.99 with free shipping.")
    }
    topic, display_order, hook = angle_map.get(angle, angle_map["整体效果"])
    price_text = f"${price:.2f}".rstrip("0").rstrip(".")
    return {
        "hook": hook,
        "selling_points": f"{topic}, {color}, beginner friendly, {price_text} free shipping",
        "display_order": display_order,
        "voiceover": (
            f"If you want an easy everyday wig, this {product_name} is the one. "
            f"Show the {topic}, check the natural movement, and it is only {price_text} with free shipping. "
            "Tap the product link before it sells out."
        ),
        "short_voiceover": (
            f"This {color} wig is beginner friendly, easy to wear, and only {price_text} shipped. "
            "Tap to get yours."
        ),
        "cover_text": f"{color} wig {price_text} shipped",
        "caption": f"{color} wig, beginner friendly, {price_text} free shipping. Tap the product link.",
        "hashtags": "#wig #syntheticwig #beginnerwig #tiktokshop #affordablewig",
        "live_pitch": f"This is {sku}, a {color} wig that is easy for beginners. It is {price_text} with free shipping today.",
        "pinned_comment": f"SKU {sku}: {color}, beginner friendly, {price_text} free shipping."
    }


def generate_tiktok_schedule(sku_rows, accounts, start_date, days=7, max_per_sku_per_day=3):
    if not sku_rows:
        return []
    parsed_start = datetime.strptime(start_date, "%Y-%m-%d")
    accounts = [a.strip() for a in accounts if a and a.strip()] or TIKTOK_DEFAULT_ACCOUNTS
    days = max(1, min(safe_int(days, 7), 30))
    max_per_sku_per_day = max(1, safe_int(max_per_sku_per_day, 3))
    last_by_account = {}
    rows = []
    sku_pool = [dict(r) for r in sku_rows]

    for day_index in range(days):
        day = (parsed_start + timedelta(days=day_index)).strftime("%Y-%m-%d")
        daily_counts = {}
        for account_index, account in enumerate(accounts):
            chosen = None
            for offset in range(len(sku_pool)):
                candidate = sku_pool[(day_index * len(accounts) + account_index + offset) % len(sku_pool)]
                sku = candidate["sku"]
                if last_by_account.get(account) == sku and len(sku_pool) > 1:
                    continue
                if daily_counts.get(sku, 0) >= max_per_sku_per_day:
                    continue
                chosen = candidate
                break
            if not chosen:
                chosen = sku_pool[(day_index + account_index) % len(sku_pool)]
            sku = chosen["sku"]
            daily_counts[sku] = daily_counts.get(sku, 0) + 1
            last_by_account[account] = sku
            angle = TIKTOK_ANGLE_POOL[(day_index + account_index) % len(TIKTOK_ANGLE_POOL)]
            script = build_tiktok_script(chosen, angle)
            rows.append({
                "publish_date": day,
                "account_name": account,
                "sku": sku,
                "video_angle": angle,
                **script
            })
    return rows


def xlsx_col_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    idx = 0
    for ch in letters:
        idx = idx * 26 + (ord(ch.upper()) - ord("A") + 1)
    return max(idx - 1, 0)


def read_xlsx_first_sheet(file_bytes: bytes):
    """Minimal XLSX reader for TikTok exports; avoids adding runtime dependencies."""
    import xml.etree.ElementTree as ET
    ns_main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    ns_rel = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    rel_ns = "{http://schemas.openxmlformats.org/package/2006/relationships}"
    with zipfile.ZipFile(BytesIO(file_bytes)) as zf:
        shared = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall(f"{ns_main}si"):
                parts = []
                for t in si.iter(f"{ns_main}t"):
                    parts.append(t.text or "")
                shared.append("".join(parts))

        sheet_path = "xl/worksheets/sheet1.xml"
        if "xl/workbook.xml" in zf.namelist() and "xl/_rels/workbook.xml.rels" in zf.namelist():
            workbook = ET.fromstring(zf.read("xl/workbook.xml"))
            rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            first_sheet = workbook.find(f".//{ns_main}sheet")
            if first_sheet is not None:
                rid = first_sheet.attrib.get(f"{ns_rel}id")
                for rel in rels.findall(f"{rel_ns}Relationship"):
                    if rel.attrib.get("Id") == rid:
                        target = rel.attrib.get("Target", "worksheets/sheet1.xml")
                        sheet_path = "xl/" + target.lstrip("/")
                        break

        root = ET.fromstring(zf.read(sheet_path))
        rows = []
        max_cols = 0
        for row in root.findall(f".//{ns_main}row"):
            values = []
            for cell in row.findall(f"{ns_main}c"):
                idx = xlsx_col_index(cell.attrib.get("r", "A1"))
                while len(values) <= idx:
                    values.append("")
                ctype = cell.attrib.get("t", "")
                text = ""
                if ctype == "inlineStr":
                    parts = []
                    for t in cell.iter(f"{ns_main}t"):
                        parts.append(t.text or "")
                    text = "".join(parts)
                else:
                    v = cell.find(f"{ns_main}v")
                    text = v.text if v is not None and v.text is not None else ""
                    if ctype == "s" and text != "":
                        try:
                            text = shared[int(text)]
                        except Exception:
                            pass
                values[idx] = text
            max_cols = max(max_cols, len(values))
            rows.append(values)
        for row in rows:
            if len(row) < max_cols:
                row.extend([""] * (max_cols - len(row)))
        while rows and not any(str(v).strip() for v in rows[0]):
            rows.pop(0)
        if not rows:
            return [], []
        headers = [str(v).strip() for v in rows[0]]
        data_rows = []
        for row in rows[1:]:
            if any(str(v).strip() for v in row):
                data_rows.append({headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))})
        return headers, data_rows


def parse_tiktok_number(value, default=0):
    text = str(value or "").strip().replace(",", "")
    if text.endswith("%"):
        text = text[:-1]
    if text.lower() in ("", "nan", "none", "-"):
        return default
    try:
        return float(text)
    except ValueError:
        return default


def parse_tiktok_int(value, default=0):
    return int(round(parse_tiktok_number(value, default)))


def parse_tiktok_publish_date(value):
    text = str(value or "").strip()
    if not text:
        return ""
    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return text[:10].replace("/", "-")


def split_tiktok_product(raw_product):
    text = str(raw_product or "").strip()
    match = re.search(r"\((\d{8,})\)\s*$", text)
    if not match:
        return text, ""
    return text[:match.start()].strip(), match.group(1)


def infer_tiktok_video_angle(text):
    lower = (text or "").lower()
    if any(word in lower for word in ["outdoor", "light", "color", "blonde", "brown", "highlight"]):
        return "颜色自然光"
    if any(word in lower for word in ["top", "part", "hairline", "natural-looking top"]):
        return "发顶自然"
    if any(word in lower for word in ["back", "layers", "full"]):
        return "背面发量"
    if any(word in lower for word in ["beginner", "glueless", "wear & go", "easy to wear"]):
        return "新手友好"
    if "23.99" in lower or "$23" in lower:
        return "23.99包邮"
    return "整体效果"


def first_sentence(text, limit=140):
    clean = " ".join(str(text or "").split())
    if not clean:
        return ""
    for sep in [".", "!", "?", "\n"]:
        pos = clean.find(sep)
        if 10 <= pos <= limit:
            return clean[:pos + 1]
    return clean[:limit]


def tiktok_perf_from_export_row(row, mapping, source_file):
    product_name, product_id = split_tiktok_product(row.get("商品"))
    mapped = mapping.get(product_id, {}) if product_id else {}
    sku = mapped.get("sku") or (f"TT-{product_id}" if product_id else "")
    simple = mapped.get("simple_sku") or simplify_sku(sku)
    video_id = str(row.get("视频ID") or "").strip()
    import_key = f"{video_id}|{product_id or product_name[:80]}"
    views = parse_tiktok_int(row.get("VV"))
    clicks = parse_tiktok_int(row.get("商品点击次数"))
    orders = parse_tiktok_int(row.get("归因 SKU 订单数"))
    gmv = parse_tiktok_number(row.get("视频归因 GMV ($)"))
    metrics = diagnose_tiktok_video({
        "views": views,
        "product_clicks": clicks,
        "orders": orders,
        "gmv": gmv,
        "comments": row.get("诊断") or ""
    })
    return {
        "import_key": import_key,
        "creator_nickname": str(row.get("达人昵称") or "").strip(),
        "creator_id": str(row.get("达人ID") or "").strip(),
        "video_info": str(row.get("视频信息") or "").strip(),
        "video_id": video_id,
        "publish_time": str(row.get("发布时间") or "").strip(),
        "publish_date": parse_tiktok_publish_date(row.get("发布时间")),
        "product_name": product_name,
        "tiktok_product_id": product_id,
        "sku": sku,
        "simple_sku": simple,
        "vv": views,
        "likes_count": parse_tiktok_int(row.get("点赞数")),
        "comments_count": parse_tiktok_int(row.get("评论数")),
        "shares_count": parse_tiktok_int(row.get("分享数")),
        "new_followers_count": parse_tiktok_int(row.get("新增粉丝数")),
        "product_redirects": parse_tiktok_int(row.get("引流次数")),
        "product_impressions": parse_tiktok_int(row.get("商品曝光次数")),
        "product_clicks": clicks,
        "unique_customers": parse_tiktok_int(row.get("去重客户数")),
        "attributed_sku_orders": orders,
        "video_sku_orders": parse_tiktok_int(row.get("视频 SKU 订单数")),
        "indirect_sku_orders": parse_tiktok_int(row.get("视频间接 SKU 订单数")),
        "attributed_units": parse_tiktok_int(row.get("视频归因成交件数")),
        "product_units": parse_tiktok_int(row.get("视频商品成交件数")),
        "indirect_units": parse_tiktok_int(row.get("视频间接成交件数")),
        "attributed_gmv": gmv,
        "video_gmv": parse_tiktok_number(row.get("视频 GMV ($)")),
        "indirect_gmv": parse_tiktok_number(row.get("视频间接 GMV ($)")),
        "gpm": parse_tiktok_number(row.get("GPM ($)")),
        "video_ctr": parse_tiktok_number(row.get("点击率（视频）")),
        "redirect_rate": parse_tiktok_number(row.get("引流率")),
        "completion_rate": parse_tiktok_number(row.get("视频完播率")),
        "ctor_sku_orders": parse_tiktok_number(row.get("CTOR（SKU 订单）")),
        "platform_diagnosis": str(row.get("诊断") or "").strip(),
        "local_diagnosis": metrics["diagnosis"],
        "repeat_action": metrics["repeat_action"],
        "source_file": source_file,
        "video_angle": infer_tiktok_video_angle(row.get("视频信息") or product_name),
        "hook": first_sentence(row.get("视频信息")),
    }


def upsert_tiktok_video_performance(conn, item):
    columns = [
        "import_key", "creator_nickname", "creator_id", "video_info", "video_id",
        "publish_time", "product_name", "tiktok_product_id", "sku", "simple_sku",
        "vv", "likes_count", "comments_count", "shares_count", "new_followers_count",
        "product_redirects", "product_impressions", "product_clicks", "unique_customers",
        "attributed_sku_orders", "video_sku_orders", "indirect_sku_orders",
        "attributed_units", "product_units", "indirect_units", "attributed_gmv",
        "video_gmv", "indirect_gmv", "gpm", "video_ctr", "redirect_rate",
        "completion_rate", "ctor_sku_orders", "platform_diagnosis", "local_diagnosis",
        "repeat_action", "source_file"
    ]
    placeholders = ", ".join("?" for _ in columns)
    update_cols = [c for c in columns if c != "import_key"]
    updates = ", ".join(f"{c}=excluded.{c}" for c in update_cols)
    sql = (
        f"INSERT INTO tiktok_video_performance ({', '.join(columns)}) VALUES ({placeholders}) "
        f"ON CONFLICT(import_key) DO UPDATE SET {updates}, updated_at=CURRENT_TIMESTAMP"
    )
    conn.execute(sql, [item.get(c) for c in columns])


def sync_perf_to_tiktok_videos(conn, item):
    existing = conn.execute(
        "SELECT id FROM tiktok_videos WHERE tiktok_video_id=? AND tiktok_product_id=?",
        (item["video_id"], item["tiktok_product_id"])
    ).fetchone()
    data = (
        item["creator_nickname"], item["sku"], item["video_id"], item["tiktok_product_id"],
        item["product_name"], item["creator_id"], item["publish_date"], item["video_angle"],
        item["hook"], item["video_info"], first_sentence(item["video_info"], 60),
        item["video_info"], "", 1, item["vv"], item["likes_count"], item["comments_count"],
        item["shares_count"], item["product_impressions"], item["product_clicks"],
        item["attributed_sku_orders"], item["attributed_gmv"], item["video_ctr"],
        item["completion_rate"], item["gpm"], "", item["platform_diagnosis"],
        item["local_diagnosis"], item["repeat_action"], item["source_file"]
    )
    if existing:
        conn.execute("""
            UPDATE tiktok_videos SET
                account_name=?, sku=?, tiktok_video_id=?, tiktok_product_id=?,
                product_name=?, creator_id=?, publish_date=?, video_angle=?, hook=?,
                voiceover=?, cover_text=?, caption=?, hashtags=?, posted=?, views=?,
                likes_count=?, comments_count=?, shares_count=?, product_impressions=?,
                product_clicks=?, orders=?, gmv=?, video_ctr=?, completion_rate=?,
                gpm=?, comments=?, platform_diagnosis=?, diagnosis=?, repeat_action=?,
                source_file=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """, data + (existing["id"],))
    else:
        conn.execute("""
            INSERT INTO tiktok_videos
                (account_name, sku, tiktok_video_id, tiktok_product_id, product_name,
                 creator_id, publish_date, video_angle, hook, voiceover, cover_text,
                 caption, hashtags, posted, views, likes_count, comments_count,
                 shares_count, product_impressions, product_clicks, orders, gmv,
                 video_ctr, completion_rate, gpm, comments, platform_diagnosis,
                 diagnosis, repeat_action, source_file)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, data)


def ensure_tiktok_ops_sku(conn, data):
    sku = str(data.get("sku") or "").strip()
    if not sku:
        return
    simple = str(data.get("simple_sku") or "").strip() or simplify_sku(sku)
    existing = conn.execute("SELECT id FROM tiktok_skus WHERE sku=?", (sku,)).fetchone()
    if existing:
        conn.execute("""
            UPDATE tiktok_skus SET
                price=COALESCE(NULLIF(?, 0), price),
                product_cost=COALESCE(NULLIF(?, 0), product_cost),
                shipping_fee=COALESCE(NULLIF(?, 0), shipping_fee),
                platform_fee_rate=COALESCE(NULLIF(?, 0), platform_fee_rate),
                ad_cost=COALESCE(NULLIF(?, 0), ad_cost),
                refund_loss=COALESCE(NULLIF(?, 0), refund_loss),
                return_rate=COALESCE(NULLIF(?, 0), return_rate),
                updated_at=CURRENT_TIMESTAMP
            WHERE sku=?
        """, (
            safe_float(data.get("price")),
            safe_float(data.get("product_cost")),
            safe_float(data.get("shipping_fee")),
            safe_float(data.get("platform_fee_rate"), 0.06),
            safe_float(data.get("ad_cost")),
            safe_float(data.get("refund_loss")),
            safe_float(data.get("return_rate"), 0.20),
            sku
        ))
    else:
        conn.execute("""
            INSERT INTO tiktok_skus
                (sku, product_name, color, length, price, product_cost, shipping_fee,
                 platform_fee_rate, ad_cost, return_rate, refund_loss, stock,
                 daily_sales, lead_time_days, safety_stock, status, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 0, 30, 10, 'testing', ?)
        """, (
            sku,
            data.get("product_name", "") or f"TikTok SKU {simple}",
            data.get("color", "") or "",
            data.get("length", "") or "",
            safe_float(data.get("price"), 23.99),
            safe_float(data.get("product_cost")),
            safe_float(data.get("shipping_fee")),
            safe_float(data.get("platform_fee_rate"), 0.06),
            safe_float(data.get("ad_cost")),
            safe_float(data.get("return_rate"), 0.20),
            safe_float(data.get("refund_loss")),
            data.get("notes", "") or "来自 TikTok SKUs 表映射"
        ))


def get_tiktok_sku_options(conn):
    """Merged SKU source for Wig Ops: ops SKUs + TikTok product mappings."""
    ops_rows = conn.execute("SELECT * FROM tiktok_skus").fetchall()
    options = {}
    for row in ops_rows:
        item = dict(row)
        sku = item.get("sku") or ""
        if not sku:
            continue
        simple = simplify_sku(sku)
        item.update({
            "simple_sku": simple,
            "tiktok_product_id": "",
            "source": "ops",
            "label": f"{sku} ({simple})",
            "search_text": f"{sku} {simple} {item.get('product_name') or ''} {item.get('color') or ''}"
        })
        options[sku] = item

    map_rows = conn.execute("SELECT * FROM tiktok_sku_mapping ORDER BY COALESCE(simple_sku, sku), sku").fetchall()
    product_names = {}
    perf_rows = conn.execute("""
        SELECT tiktok_product_id, product_name
        FROM tiktok_video_performance
        WHERE tiktok_product_id IS NOT NULL AND tiktok_product_id != ''
        GROUP BY tiktok_product_id
    """).fetchall()
    for row in perf_rows:
        product_names[str(row["tiktok_product_id"])] = row["product_name"] or ""

    for row in map_rows:
        m = dict(row)
        sku = m.get("sku") or ""
        if not sku:
            continue
        simple = m.get("simple_sku") or simplify_sku(sku)
        product_id = str(m.get("tiktok_product_id") or "")
        product_name = product_names.get(product_id, "")
        if sku in options:
            item = options[sku]
            item["simple_sku"] = item.get("simple_sku") or simple
            item["tiktok_product_id"] = product_id or item.get("tiktok_product_id", "")
            item["source"] = "ops+mapping"
            if product_name and (not item.get("product_name") or str(item.get("product_name")).startswith("TikTok SKU ")):
                item["product_name"] = product_name
            for key in ["price", "product_cost", "shipping_fee", "platform_fee_rate", "ad_cost", "refund_loss", "return_rate"]:
                if safe_float(item.get(key)) == 0 and safe_float(m.get(key)) != 0:
                    item[key] = m.get(key)
        else:
            item = {
                "id": None,
                "sku": sku,
                "simple_sku": simple,
                "tiktok_product_id": product_id,
                "product_name": product_name or f"TikTok SKU {simple}",
                "color": "",
                "length": "",
                "price": safe_float(m.get("price"), 23.99),
                "product_cost": safe_float(m.get("product_cost")),
                "shipping_fee": safe_float(m.get("shipping_fee")),
                "platform_fee_rate": safe_float(m.get("platform_fee_rate"), 0.06),
                "ad_cost": safe_float(m.get("ad_cost")),
                "return_rate": safe_float(m.get("return_rate"), 0.20),
                "refund_loss": safe_float(m.get("refund_loss")),
                "stock": 0,
                "daily_sales": 0,
                "lead_time_days": 30,
                "safety_stock": 10,
                "status": "testing",
                "notes": m.get("notes") or "来自 TikTok SKUs 表映射",
                "source": "mapping",
            }
        item["label"] = f"{sku} ({simple})"
        if item.get("tiktok_product_id"):
            item["label"] += f" · TT {item['tiktok_product_id']}"
        item["search_text"] = " ".join([
            str(item.get("sku") or ""),
            str(item.get("simple_sku") or ""),
            str(item.get("tiktok_product_id") or ""),
            str(item.get("product_name") or ""),
            str(item.get("color") or "")
        ])
        options[sku] = item
    return list(options.values())


def find_tiktok_sku_info(conn, token):
    token = str(token or "").strip()
    if not token:
        return None
    lowered = token.lower()
    options = get_tiktok_sku_options(conn)
    for item in options:
        candidates = [
            str(item.get("sku") or ""),
            str(item.get("simple_sku") or ""),
            str(item.get("tiktok_product_id") or ""),
            str(item.get("label") or "")
        ]
        if any(lowered == c.lower() for c in candidates if c):
            return item
    for item in options:
        if lowered in str(item.get("search_text") or "").lower():
            return item
    return None


# ────────────────────────────────────────
# Shopify 连接器
# ────────────────────────────────────────
class ShopifyConnector:
    def __init__(self, shop: str, access_token: str = "", api_key: str = "", api_password: str = ""):
        self.shop = shop
        self.base_url = f"https://{shop}/admin/api/2025-04"
        self.api_key = api_key
        self.api_password = api_password
        self.headers = {"Content-Type": "application/json"}
        self._set_access_token(access_token)

    def _set_access_token(self, token: str):
        if token and len(token) > 10:
            self.headers["X-Shopify-Access-Token"] = token
            self.access_token = token
        elif self.api_key and self.api_password:
            auth = base64.b64encode(f"{self.api_key}:{self.api_password}".encode()).decode()
            self.headers["Authorization"] = f"Basic {auth}"
        else:
            raise ValueError("Need SHOPIFY_ACCESS_TOKEN or SHOPIFY_API_KEY + SHOPIFY_API_PASSWORD")

    async def refresh_token(self):
        """刷新 Shopify access token (24小时过期)"""
        import httpx
        async with httpx.AsyncClient() as client:
            resp = await client.post(
                f"https://{self.shop}/admin/oauth/access_token",
                data={
                    "grant_type": "client_credentials",
                    "client_id": self.api_key,
                    "client_secret": self.api_password
                }
            )
            if resp.status_code == 200:
                data = resp.json()
                new_token = data.get("access_token")
                if new_token:
                    self._set_access_token(new_token)
                    # 更新 .env
                    env_file = BASE_DIR / ".env"
                    if env_file.exists():
                        content = env_file.read_text()
                        content = __import__('re').sub(
                            r'SHOPIFY_ACCESS_TOKEN=.*',
                            f'SHOPIFY_ACCESS_TOKEN={new_token}',
                            content
                        )
                        env_file.write_text(content)
                    return new_token
            return None

    async def get_orders(self, limit=50, status="any", since_id=None):
        """拉取订单列表"""
        url = f"{self.base_url}/orders.json"
        params = {"limit": limit, "status": status}
        if since_id:
            params["since_id"] = since_id
        async with httpx.AsyncClient() as client:
            resp = await client.get(url, headers=self.headers, params=params)
            resp.raise_for_status()
            return resp.json().get("orders", [])

    async def get_products(self, limit=50):
        """拉取产品列表"""
        url = f"{self.base_url}/products.json"
        params = {"limit": limit}
        async with httpx.AsyncClient() as client:
            resp = await client.get(url, headers=self.headers, params=params)
            resp.raise_for_status()
            return resp.json().get("products", [])

    async def get(self, path: str):
        """通用 GET 请求"""
        import httpx
        url = f"{self.base_url}{path}"
        async with httpx.AsyncClient() as client:
            resp = await client.get(url, headers=self.headers)
            resp.raise_for_status()
            return resp.json()

    async def delete(self, path: str):
        """通用 DELETE 请求"""
        url = f"{self.base_url}{path}"
        async with httpx.AsyncClient() as client:
            resp = await client.delete(url, headers=self.headers)
            resp.raise_for_status()
            return resp.json()

    async def set_inventory(self, inventory_item_id: int, location_id: int, available: int, variant_id: int = None):
        """设置Shopify库存数量（自动开启跟踪+连接）"""
        import httpx
        url = f"{self.base_url}/inventory_levels/set.json"
        async with httpx.AsyncClient() as client:
            resp = await client.post(url, headers=self.headers, json={
                "location_id": location_id,
                "inventory_item_id": inventory_item_id,
                "available": available
            })
            # 422 常见原因：未跟踪库存 / 库存项未启用
            if resp.status_code == 422:
                body_text = resp.text
                # 1) 变体未开启库存跟踪 → 更新variant启用shopify库存管理
                if variant_id and "does not have inventory" in body_text.lower():
                    var_url = f"{self.base_url}/variants/{variant_id}.json"
                    await client.put(var_url, headers=self.headers, json={
                        "variant": {"id": variant_id, "inventory_management": "shopify"}
                    })
                # 2) 库存项未连接到此位置 → 先连接
                connect_url = f"{self.base_url}/inventory_levels/connect.json"
                resp2 = await client.post(connect_url, headers=self.headers, json={
                    "location_id": location_id,
                    "inventory_item_id": inventory_item_id,
                    "relocate_if_necessary": True
                })
                if resp2.status_code in (200, 201):
                    resp = await client.post(url, headers=self.headers, json={
                        "location_id": location_id,
                        "inventory_item_id": inventory_item_id,
                        "available": available
                    })
                # 如果依然422，记录详细错误
                if resp.status_code == 422:
                    raise Exception(f"422: {resp.text[:150]}")
            resp.raise_for_status()
            return resp.json()

    async def untrack_variant(self, variant_id: int):
        """取消变体库存跟踪（允许无限售卖）"""
        import httpx
        url = f"{self.base_url}/variants/{variant_id}.json"
        async with httpx.AsyncClient() as client:
            resp = await client.put(url, headers=self.headers, json={
                "variant": {"id": variant_id, "inventory_management": None}
            })
            resp.raise_for_status()
            return resp.json()

    async def get_inventory_levels(self, inventory_item_ids: list):
        """查询库存"""
        url = f"{self.base_url}/inventory_levels.json"
        params = {"inventory_item_ids": ",".join(map(str, inventory_item_ids))}
        async with httpx.AsyncClient() as client:
            resp = await client.get(url, headers=self.headers, params=params)
            resp.raise_for_status()
            return resp.json().get("inventory_levels", [])

    async def get_order_fulfillments(self, order_id: int):
        """获取订单的履约记录（含运单号）"""
        url = f"{self.base_url}/orders/{order_id}/fulfillments.json"
        async with httpx.AsyncClient() as client:
            resp = await client.get(url, headers=self.headers)
            resp.raise_for_status()
            return resp.json().get("fulfillments", [])

    async def create_fulfillment(self, order_id: int, tracking_number: str,
                                  tracking_company: str = "云途物流",
                                  line_items: list = None):
        """创建发货（回传运单到 Shopify）"""
        url = f"{self.base_url}/orders/{order_id}/fulfillments.json"
        payload = {
            "fulfillment": {
                "tracking_number": tracking_number,
                "tracking_company": tracking_company,
                "notify_customer": True
            }
        }
        if line_items:
            payload["fulfillment"]["line_items"] = line_items

        async with httpx.AsyncClient() as client:
            resp = await client.post(url, headers=self.headers, json=payload)
            resp.raise_for_status()
            return resp.json()

    def verify_webhook(self, data: bytes, hmac_header: str) -> bool:
        """验证 Shopify Webhook HMAC 签名"""
        if not SHOPIFY_WEBHOOK_SECRET:
            return True  # 开发模式跳过验证
        digest = hmac.new(
            SHOPIFY_WEBHOOK_SECRET.encode(),
            data,
            hashlib.sha256
        ).digest()
        computed = base64.b64encode(digest).decode()
        return hmac.compare_digest(computed, hmac_header)


# ────────────────────────────────────────
# FastAPI 应用
# ────────────────────────────────────────
async def token_refresh_loop():
    """每12小时刷新一次 Shopify Token"""
    # 首次启动立即刷新
    try:
        new_token = await shopify.refresh_token()
        if new_token:
            print(f"[Token] Startup refresh: {new_token[:20]}...")
    except Exception as e:
        print(f"[Token] Startup refresh failed: {e}")
    while True:
        await asyncio.sleep(43200)  # 12小时
        try:
            new_token = await shopify.refresh_token()
            if new_token:
                print(f"[Token] Refreshed: {new_token[:20]}...")
        except Exception as e:
            print(f"[Token] Refresh failed: {e}")

@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    asyncio.create_task(token_refresh_loop())
    # 启动邮件提醒后台线程
    remind_thread = threading.Thread(target=reminder_loop, daemon=True)
    remind_thread.start()
    yield


def _next_remind_date(current_date: str, repeat_type: str, repeat_day: int) -> str:
    """计算下一次提醒日期"""
    from datetime import timedelta
    dt = datetime.strptime(current_date, "%Y-%m-%d")
    if repeat_type == "monthly":
        # 下个月的同一天（处理月末溢出）
        month = dt.month + 1
        year = dt.year
        if month > 12:
            month = 1
            year += 1
        import calendar
        max_day = calendar.monthrange(year, month)[1]
        day = min(repeat_day, max_day)
        return f"{year}-{month:02d}-{day:02d}"
    elif repeat_type == "weekly":
        # 下周同一天
        next_dt = dt + timedelta(days=7)
        return next_dt.strftime("%Y-%m-%d")
    return current_date

def reminder_loop():
    """后台线程: 每分钟检查是否有到期的提醒"""
    while True:
        try:
            conn = sqlite3.connect(str(DB_PATH))
            conn.row_factory = sqlite3.Row
            now = datetime.now().strftime("%Y-%m-%d")
            rows = conn.execute(
                "SELECT id, title, content, remind_date, email, repeat_type, repeat_day"
                " FROM reminders WHERE remind_date <= ? AND sent = 0",
                (now,)
            ).fetchall()
            for r in rows:
                repeat_label = ""
                if r["repeat_type"] == "monthly":
                    repeat_label = f" (每月{r['repeat_day']}号重复)"
                elif r["repeat_type"] == "weekly":
                    weekday_names = ["一","二","三","四","五","六","日"]
                    wd = weekday_names[r["repeat_day"]] if 0 <= r["repeat_day"] <= 6 else ""
                    repeat_label = f" (每周{wd}重复)" if wd else " (每周重复)"

                send_email(
                    to=r["email"] or EMAIL_TO,
                    subject=f"⏰ 提醒: {r['title']}{repeat_label}",
                    body=f"📌 {r['title']}{repeat_label}\n\n"
                         f"本次提醒日期：{r['remind_date']}\n\n"
                         f"{r['content'] or '(无详细内容)'}\n\n"
                         + (f"🔄 重复提醒，下次自动推送到下一周期。\n" if r["repeat_type"] else "")
                         + f"—— LuxeLocks Hub 自动发送"
                )
                repeat_type = r["repeat_type"] or ""
                if repeat_type in ("monthly", "weekly"):
                    # 重复提醒：计算下一次日期，保持 sent=0
                    next_date = _next_remind_date(r["remind_date"], repeat_type, r["repeat_day"])
                    conn.execute(
                        "UPDATE reminders SET remind_date = ? WHERE id = ?",
                        (next_date, r["id"])
                    )
                else:
                    # 一次性提醒：标记已发送
                    conn.execute("UPDATE reminders SET sent = 1 WHERE id = ?", (r["id"],))
                conn.commit()
            conn.close()
        except Exception as e:
            pass  # 静默重试
        time.sleep(60)


def send_email(to: str, subject: str, body: str):
    """通过QQ邮箱SMTP发送邮件"""
    msg = MIMEMultipart()
    msg["From"] = EMAIL_USER
    msg["To"] = to
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))
    try:
        server = smtplib.SMTP(EMAIL_HOST, EMAIL_PORT, timeout=10)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASS)
        server.sendmail(EMAIL_USER, [to], msg.as_string())
        server.quit()
        return True
    except Exception:
        return False

app = FastAPI(
    title="LuxeLocks Hub",
    description="轻量跨境电商订单管理中台",
    version="1.0.0",
    lifespan=lifespan
)

jinja_env = Jinja2Templates(directory=str(BASE_DIR / "templates")).env

def render_html(name: str, request: Request, **context) -> HTMLResponse:
    """渲染模板（绕过 Jinja2 缓存 bug）"""
    tmpl = jinja_env.get_template(name)
    return HTMLResponse(tmpl.render({"request": request, **context}))

# 全局 Shopify 连接器实例
shopify = ShopifyConnector(
    SHOPIFY_SHOP,
    access_token=SHOPIFY_ACCESS_TOKEN,
    api_key=SHOPIFY_API_KEY,
    api_password=SHOPIFY_API_PASSWORD
)


# ────────────────────────────────────────
# Webhook 接收
# ────────────────────────────────────────
@app.post("/webhook/shopify/order")
async def shopify_order_webhook(request: Request):
    """接收 Shopify 订单 Webhook"""
    body = await request.body()
    data = json.loads(body)

    conn = get_db()
    try:
        order = data
        line_items = json.dumps([
            {
                "title": item.get("title", ""),
                "sku": item.get("sku", ""),
                "quantity": item.get("quantity", 0),
                "price": item.get("price", "0")
            }
            for item in order.get("line_items", [])
        ]) if order.get("line_items") else "[]"

        shipping = order.get("shipping_address", {}) or order.get("billing_address", {}) or {}
        shipping_str = json.dumps(shipping)

        # 使用 Shopify 的订单创建时间
        order_created = order.get("created_at", datetime.now().isoformat())
        conn.execute("""
            INSERT OR REPLACE INTO orders
            (platform, platform_order_id, order_number, customer_name, customer_email,
             total_price, currency, status, shipping_address, line_items, raw_data, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        """, (
            "shopify",
            str(order.get("id", "")),
            order.get("name", order.get("order_number", "")),
            f"{shipping.get('first_name', '')} {shipping.get('last_name', '')}".strip(),
            order.get("email", order.get("contact_email", "")),
            float(order.get("total_price", order.get("current_total_price", 0))),
            order.get("currency", "USD"),
            order.get("financial_status", "pending"),
            shipping_str,
            line_items,
            json.dumps(data, ensure_ascii=False),
            order_created
        ))
        conn.commit()

        # 记录同步日志
        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("shopify", "webhook_order", "success", f"Order #{order.get('name')}")
        )
        conn.commit()

        return {"status": "ok"}
    except Exception as e:
        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("shopify", "webhook_order", "error", str(e))
        )
        conn.commit()
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/webhook/shopify/fulfillment")
async def shopify_fulfillment_webhook(request: Request):
    """接收 Shopify 发货更新 Webhook"""
    body = await request.body()
    data = json.loads(body)

    conn = get_db()
    try:
        order_id = str(data.get("order_id", ""))
        tracking_number = data.get("tracking_number", "")
        tracking_company = data.get("tracking_company", "")

        conn.execute("""
            UPDATE orders SET tracking_number = ?, tracking_company = ?,
            status = 'shipped', updated_at = CURRENT_TIMESTAMP
            WHERE platform_order_id = ?
        """, (tracking_number, tracking_company, order_id))
        conn.commit()

        return {"status": "ok"}
    finally:
        conn.close()


# ────────────────────────────────────────
# 网页看板
# ────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request):
    """订单管理主页"""
    conn = get_db()
    try:
        # 统计
        total_orders = conn.execute("SELECT COUNT(*) FROM orders").fetchone()[0]
        pending = conn.execute(
            "SELECT COUNT(*) FROM orders WHERE status IN ('pending','paid','authorized')"
        ).fetchone()[0]
        shipped = conn.execute("SELECT COUNT(*) FROM orders WHERE status = 'shipped'").fetchone()[0]
        today = conn.execute(
            "SELECT COUNT(*) FROM orders WHERE date(created_at) = date('now')"
        ).fetchone()[0]

        # 最近订单
        orders = conn.execute("""
            SELECT * FROM orders ORDER BY created_at DESC LIMIT 50
        """).fetchall()

        # 低库存产品
        low_stock = conn.execute("""
            SELECT * FROM products WHERE inventory_quantity < 10
            ORDER BY inventory_quantity ASC LIMIT 20
        """).fetchall()

        return render_html("dashboard.html", request,
            stats={
                "total": total_orders,
                "pending": pending,
                "shipped": shipped,
                "today": today
            },
            orders=[dict(o) for o in orders],
            low_stock=[dict(p) for p in low_stock]
        )
    finally:
        conn.close()


@app.get("/orders", response_class=HTMLResponse)
async def order_list(request: Request, status: str = None, search: str = None):
    """订单列表页"""
    conn = get_db()
    try:
        query = "SELECT * FROM orders WHERE 1=1"
        params = []
        if status:
            query += " AND status = ?"
            params.append(status)
        if search:
            query += " AND (order_number LIKE ? OR platform_order_id LIKE ? OR customer_name LIKE ?)"
            search_term = f"%{search}%"
            params.extend([search_term, search_term, search_term])
        query += " ORDER BY created_at DESC LIMIT 200"

        orders = conn.execute(query, params).fetchall()
        return render_html("orders.html", request,
            orders=[dict(o) for o in orders],
            status_filter=status,
            search=search
        )
    finally:
        conn.close()


@app.get("/order/{order_id}", response_class=HTMLResponse)
async def order_detail(request: Request, order_id: str):
    """订单详情"""
    conn = get_db()
    try:
        order = conn.execute(
            "SELECT * FROM orders WHERE platform_order_id = ? OR id = ?",
            (order_id, order_id if order_id.isdigit() else 0)
        ).fetchone()
        if not order:
            raise HTTPException(404, "订单不存在")

        items = json.loads(order["line_items"] or "[]")

        # ── 匹配采购资源信息池 ──
        # 收集所有简化SKU
        simple_skus = []
        for item in items:
            ss = simplify_sku(item.get("sku", ""))
            if ss:
                simple_skus.append(ss)

        procurement_items = []
        if simple_skus:
            # 批量查询采购资源
            placeholders = ",".join(["?" for _ in simple_skus])
            proc_rows = conn.execute(
                f"SELECT * FROM procurement_resources WHERE simple_sku IN ({placeholders})",
                simple_skus
            ).fetchall()
            proc_map = {r["simple_sku"]: dict(r) for r in proc_rows}

            # 加载供应商数据
            sup_rows = conn.execute(
                "SELECT name, seller_id, purchase_link, contact_name, wechat FROM suppliers"
            ).fetchall()
            sup_map = {}
            for s in sup_rows:
                sid = (s["seller_id"] or "").strip()
                if sid:
                    sup_map[sid] = {
                        "name": s["name"],
                        "purchase_link": s["purchase_link"] or "",
                        "contact_name": s["contact_name"] or "",
                        "wechat": s["wechat"] or ""
                    }

            # 为每个SKU组装采购明细
            for item in items:
                ss = simplify_sku(item.get("sku", ""))
                if not ss:
                    continue
                proc = proc_map.get(ss)
                if not proc:
                    continue

                # 供应商数据（从关联的seller_id获取）
                seller_id = (proc.get("seller_id") or "").strip()
                sup = sup_map.get(seller_id, {})
                procurement_items.append({
                    "simple_sku": ss,
                    "full_sku": item.get("sku", ""),
                    "supplier": proc.get("supplier") or sup.get("name", ""),
                    "seller_id": seller_id or proc.get("seller_id", ""),
                    "cost": proc.get("cost") or 0,
                    "purchase_link": proc.get("purchase_link") or sup.get("purchase_link", ""),
                    "contact_name": proc.get("contact_name") or sup.get("contact_name", ""),
                    "wechat": proc.get("wechat") or sup.get("wechat", ""),
                })

        return render_html("order_detail.html", request,
            order=dict(order),
            line_items=items,
            shipping=json.loads(order["shipping_address"] or "{}"),
            procurement_items=procurement_items
        )
    finally:
        conn.close()


@app.get("/api/search")
async def api_search(q: str = Query(...)):
    """搜索订单（API）"""
    conn = get_db()
    try:
        results = conn.execute("""
            SELECT * FROM orders
            WHERE order_number LIKE ? OR platform_order_id LIKE ?
               OR customer_name LIKE ? OR tracking_number LIKE ?
            ORDER BY created_at DESC LIMIT 20
        """, (f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%")).fetchall()
        return {"results": [dict(r) for r in results]}
    finally:
        conn.close()


# ────────────────────────────────────────
# 手动同步
# ────────────────────────────────────────
@app.post("/api/sync/orders")
async def sync_orders():
    """手动同步 Shopify 订单"""
    conn = get_db()
    try:
        # ── 预加载跨境库存用于计算运输类型 ──
        wh_items = conn.execute(
            "SELECT reference_code, available_inventory FROM warehouse_inventory"
            " WHERE warehouse_name IN ('CrossBorder', 'NewProducts')"
        ).fetchall()
        cross_stock = {}
        for w in wh_items:
            ref = simplify_sku(w["reference_code"])
            if ref:
                cross_stock[ref] = cross_stock.get(ref, 0) + w["available_inventory"]

        mappings = conn.execute(
            "SELECT shopify_simple_sku, cross_border_simple_sku FROM sku_mapping"
        ).fetchall()
        map_cross = {}
        for m in mappings:
            if m["cross_border_simple_sku"]:
                map_cross[m["shopify_simple_sku"]] = m["cross_border_simple_sku"]

        orders = await shopify.get_orders(limit=50)
        count = 0
        for order in orders:
            line_items_raw = [
                {
                    "title": item.get("title", ""),
                    "sku": item.get("sku", ""),
                    "quantity": item.get("quantity", 0),
                    "price": item.get("price", "0")
                }
                for item in order.get("line_items", [])
            ] if order.get("line_items") else []
            line_items = json.dumps(line_items_raw)

            shipping = order.get("shipping_address", {}) or {}
            shipping_str = json.dumps(shipping)
            order_created = order.get("created_at", datetime.now().isoformat())
            new_status = order.get("financial_status", "pending")
            platform_id = str(order.get("id", ""))

            # ── 计算运输类型 ──
            # 若订单已存在且为已发货状态，保留原有运输类型
            old_shipping = conn.execute(
                "SELECT shipping_type, status FROM orders WHERE platform_order_id = ? AND platform = 'shopify'",
                (platform_id,)
            ).fetchone()
            if old_shipping and old_shipping["status"] == "shipped" and old_shipping["shipping_type"]:
                computed_type = old_shipping["shipping_type"]
            else:
                computed_type = "国内直发"
                for item in line_items_raw:
                    ss = simplify_sku(item.get("sku", ""))
                    cross_key = map_cross.get(ss, ss)
                    if cross_stock.get(cross_key, 0) > 0:
                        computed_type = "海外仓发货"
                        break

            conn.execute("""
                INSERT OR REPLACE INTO orders
                (platform, platform_order_id, order_number, customer_name, customer_email,
                 total_price, currency, status, shipping_address, line_items, raw_data,
                 shipping_type, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            """, (
                "shopify",
                platform_id,
                order.get("name", ""),
                f"{shipping.get('first_name','')} {shipping.get('last_name','')}".strip(),
                order.get("email", ""),
                float(order.get("total_price", 0)),
                order.get("currency", "USD"),
                new_status,
                shipping_str,
                line_items,
                json.dumps(order, ensure_ascii=False),
                computed_type,
                order_created
            ))
            count += 1

        conn.commit()
        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("shopify", "manual_sync", "success", f"Synced {count} orders")
        )
        conn.commit()
        return {"status": "ok", "synced": count}
    except Exception as e:
        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("shopify", "manual_sync", "error", str(e))
        )
        conn.commit()
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/api/sync/tracking")
async def sync_tracking():
    """从 Shopify 同步运单号（OMS 回传到 Shopify 的履约数据）"""
    conn = get_db()
    try:
        # 查找所有未发货的订单
        orders = conn.execute(
            "SELECT platform_order_id, order_number FROM orders WHERE status != 'shipped' AND status != 'cancelled'"
        ).fetchall()

        updated = 0
        for order in orders:
            oid = order["platform_order_id"]
            try:
                fulfillments = await shopify.get_order_fulfillments(int(oid))
                for f in fulfillments:
                    tracking_number = f.get("tracking_number", "")
                    tracking_company = f.get("tracking_company", "OMS")
                    if tracking_number:
                        conn.execute("""
                            UPDATE orders SET tracking_number = ?, tracking_company = ?,
                            status = 'shipped', updated_at = CURRENT_TIMESTAMP
                            WHERE platform_order_id = ?
                        """, (tracking_number, tracking_company, oid))
                        updated += 1
                        # 只需一个履约记录
                        break
            except Exception:
                pass

        conn.commit()
        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("shopify", "sync_tracking", "success", f"Updated {updated} orders with tracking")
        )
        conn.commit()
        return {"status": "ok", "updated": updated}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/api/sync/products")
async def sync_products():
    """手动同步 Shopify 产品（增量更新 + 删除 Shopify 已不存在的产品）"""
    conn = get_db()
    try:
        products = await shopify.get_products(limit=250)
        shopify_skus = set()
        count = 0
        for product in products:
            pid = str(product.get("id", ""))
            # 构建 product images 索引：image_id → src
            product_images = product.get("images", []) or []
            image_map = {}
            for img in product_images:
                img_id = img.get("id")
                img_src = img.get("src", "")
                if img_id:
                    image_map[str(img_id)] = img_src
            # 产品级主图作为兜底
            fallback_img = (product.get("image", {}) or {}).get("src", "")
            if not fallback_img and product_images:
                fallback_img = product_images[0].get("src", "")

            for variant in product.get("variants", []):
                sku = variant.get("sku", "")
                if not sku:
                    continue
                shopify_skus.add(sku)

                # 优先使用 variant 专属图片，否则用产品主图
                variant_image_id = variant.get("image_id")
                img_url = ""
                if variant_image_id and str(variant_image_id) in image_map:
                    img_url = image_map[str(variant_image_id)]
                if not img_url:
                    img_url = fallback_img

                conn.execute("""
                    INSERT OR REPLACE INTO products
                    (platform, platform_product_id, sku, title, variant_title,
                     price, inventory_quantity, image_url, product_status,
                     variant_id, inventory_item_id, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                """, (
                    "shopify",
                    pid,
                    sku,
                    product.get("title", ""),
                    variant.get("title", ""),
                    float(variant.get("price", "0")),
                    int(variant.get("inventory_quantity", 0)),
                    img_url,
                    product.get("status", "active"),
                    str(variant.get("id", "")),
                    str(variant.get("inventory_item_id", ""))
                ))
                count += 1

        # 删除 Shopify 中已不存在的产品（如 2088-1）
        db_rows = conn.execute(
            "SELECT sku FROM products WHERE platform = 'shopify'"
        ).fetchall()
        deleted_count = 0
        for row in db_rows:
            db_sku = row["sku"]
            if db_sku and db_sku not in shopify_skus:
                conn.execute(
                    "DELETE FROM products WHERE platform = 'shopify' AND sku = ?",
                    (db_sku,)
                )
                deleted_count += 1

        conn.commit()
        detail = f"Synced {count} variants, deleted {deleted_count} stale"
        if deleted_count > 0:
            detail += f" (removed: {deleted_count} products no longer in Shopify)"
        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("shopify", "manual_sync_products", "success", detail)
        )
        conn.commit()
        return {"status": "ok", "synced": count, "deleted": deleted_count}
    except Exception as e:
        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("shopify", "manual_sync_products", "error", str(e))
        )
        conn.commit()
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/api/products/clear")
async def products_clear():
    """清空本地所有产品数据"""
    conn = get_db()
    try:
        conn.execute("DELETE FROM products WHERE platform = 'shopify'")
        conn.commit()
        deleted = conn.total_changes
        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("shopify", "manual_clear_products", "success", f"Cleared {deleted} products from local DB")
        )
        conn.commit()
        return {"status": "ok", "deleted": deleted}
    except Exception as e:
        try: conn.rollback()
        except: pass
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/api/product/delete")
async def product_delete(request: Request):
    """删除单个产品（从本地DB + Shopify）"""
    data = await request.json()
    sku = data.get("sku", "").strip()
    if not sku:
        return {"status": "error", "message": "SKU required"}

    conn = get_db()
    try:
        # 1. 从 Shopify 删除（如果还存在）
        row = conn.execute(
            "SELECT platform_product_id, sku FROM products WHERE sku = ? AND platform = 'shopify'",
            (sku,)
        ).fetchone()
        shopify_deleted = False
        if row and row["platform_product_id"]:
            pid = row["platform_product_id"]
            try:
                await shopify.delete(f"/products/{pid}.json")
                shopify_deleted = True
            except Exception:
                shopify_deleted = False

        # 2. 从本地 DB 删除
        conn.execute("DELETE FROM products WHERE sku = ? AND platform = 'shopify'", (sku,))
        conn.commit()

        detail = f"Deleted SKU: {sku}"
        if shopify_deleted:
            detail += " (also from Shopify)"
        else:
            detail += " (local DB only — already gone from Shopify)"
        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("shopify", "manual_delete_product", "success", detail)
        )
        conn.commit()
        return {"status": "ok", "sku": sku, "shopify_deleted": shopify_deleted}
    except Exception as e:
        try: conn.rollback()
        except: pass
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


# ────────────────────────────────────────
# MCP API 接口（给 MCP Server 调用）
# ────────────────────────────────────────
@app.get("/api/mcp/orders")
async def mcp_get_orders(
    status: str = None,
    limit: int = 20,
    search: str = None
):
    """MCP: 查询订单"""
    conn = get_db()
    try:
        query = "SELECT * FROM orders WHERE 1=1"
        params = []
        if status:
            query += " AND status = ?"
            params.append(status)
        if search:
            q = f"%{search}%"
            query += " AND (order_number LIKE ? OR customer_name LIKE ? OR tracking_number LIKE ?)"
            params.extend([q, q, q])
        query += " ORDER BY created_at DESC LIMIT ?"
        params.append(limit)

        orders = conn.execute(query, params).fetchall()
        return {"orders": [dict(o) for o in orders], "count": len(orders)}
    finally:
        conn.close()


@app.get("/api/mcp/inventory")
async def mcp_get_inventory(sku: str = None):
    """MCP: 查询库存"""
    conn = get_db()
    try:
        query = "SELECT * FROM products WHERE 1=1"
        params = []
        if sku:
            query += " AND sku LIKE ?"
            params.append(f"%{sku}%")
        query += " ORDER BY inventory_quantity ASC LIMIT 50"
        products = conn.execute(query, params).fetchall()
        return {"products": [dict(p) for p in products]}
    finally:
        conn.close()


@app.get("/api/mcp/stats")
async def mcp_get_stats():
    """MCP: 今日/总计汇总"""
    conn = get_db()
    try:
        today_orders = conn.execute(
            "SELECT COUNT(*) FROM orders WHERE date(created_at) = date('now')"
        ).fetchone()[0]
        pending = conn.execute(
            "SELECT COUNT(*) FROM orders WHERE status IN ('pending','paid','authorized')"
        ).fetchone()[0]
        shipped_today = conn.execute(
            "SELECT COUNT(*) FROM orders WHERE status = 'shipped' AND date(updated_at) = date('now')"
        ).fetchone()[0]
        low_stock = conn.execute(
            "SELECT COUNT(*) FROM products WHERE inventory_quantity < 10 AND inventory_quantity >= 0"
        ).fetchone()[0]
        return {
            "today_orders": today_orders,
            "pending_orders": pending,
            "shipped_today": shipped_today,
            "low_stock_count": low_stock
        }
    finally:
        conn.close()


@app.get("/api/mcp/sync_log")
async def mcp_get_sync_log(limit: int = 20):
    """MCP: 查看同步日志"""
    conn = get_db()
    try:
        logs = conn.execute(
            "SELECT * FROM sync_log ORDER BY created_at DESC LIMIT ?", (limit,)
        ).fetchall()
        return {"logs": [dict(l) for l in logs]}
    finally:
        conn.close()


# ────────────────────────────────────────
# 产品图片上传
# ────────────────────────────────────────
@app.post("/api/product/image")
async def product_upload_image(request: Request):
    """更新产品图片URL"""
    data = await request.json()
    sku = data.get("sku", "")
    pid = data.get("platform_product_id", "")
    image_url = data.get("image_url", "")

    if not image_url:
        return {"status": "error", "message": "请提供图片URL"}

    conn = get_db()
    try:
        conn.execute(
            "UPDATE products SET image_url = ?, updated_at = CURRENT_TIMESTAMP WHERE sku = ? OR platform_product_id = ?",
            (image_url, sku, pid)
        )
        if conn.total_changes == 0:
            return {"status": "error", "message": "未找到该产品"}
        conn.commit()
        return {"status": "ok", "message": "图片已更新"}
    finally:
        conn.close()


# ────────────────────────────────────────
# 产品管理页面
# ────────────────────────────────────────
def simplify_sku(sku: str) -> str:
    """简化 SKU：从完整 SKU 提取 款式编号-变体编号"""
    import re
    if not sku:
        return ""
    # 匹配 数字-数字 模式（保留变体编号的原始格式，包括前导零）
    match = re.search(r'(\d+-\d+)', sku)
    if match:
        return match.group(1)
    return sku


@app.get("/products", response_class=HTMLResponse)
async def product_list(request: Request, search: str = None, status: str = None, sort: str = None, order: str = "asc"):
    conn = get_db()
    try:
        # 构建查询
        conditions = []
        params = []
        if search:
            search_term = search.strip()
            conditions.append("sku LIKE ?")
            params.append(f"%{search_term}%")
        if status and status != "all":
            conditions.append("product_status = ?")
            params.append(status)

        where_clause = ""
        if conditions:
            where_clause = "WHERE " + " AND ".join(conditions)

        # 默认按Shopify库存排序（SQL），其余列在Python层面排序
        query = f"SELECT * FROM products {where_clause} ORDER BY inventory_quantity DESC LIMIT 200"
        products = conn.execute(query, params).fetchall()

        # ── 加载仓库库存，按简化SKU建索引 ──
        wh_items = conn.execute(
            "SELECT reference_code, warehouse_name, available_inventory, in_transit_total FROM warehouse_inventory"
        ).fetchall()
        wh_cross = {}   # simplified_sku → {available, transit}
        wh_luxe = {}
        wh_veloura = {}
        for w in wh_items:
            ref = simplify_sku(w["reference_code"])
            if not ref:
                continue
            target = None
            if w["warehouse_name"] == "CrossBorder" or w["warehouse_name"] == "NewProducts":
                target = wh_cross
            elif w["warehouse_name"] == "LuxeLocks":
                target = wh_luxe
            elif w["warehouse_name"] == "VelouraHair":
                target = wh_veloura
            if target is not None:
                if ref in target:
                    target[ref]["available"] += w["available_inventory"]
                    target[ref]["transit"] += w["in_transit_total"]
                else:
                    target[ref] = {"available": w["available_inventory"], "transit": w["in_transit_total"]}

        # ── 加载 SKU 映射表 ──
        mappings = conn.execute("SELECT * FROM sku_mapping").fetchall()
        map_cross = {}    # shopify_simple → cross_border_simple
        map_luxe = {}     # shopify_simple → luxelocks_simple
        map_veloura = {}  # shopify_simple → velourahair_simple
        for m in mappings:
            ss = m["shopify_simple_sku"]
            if m["cross_border_simple_sku"]:
                map_cross[ss] = m["cross_border_simple_sku"]
            if m["luxelocks_simple_sku"]:
                map_luxe[ss] = m["luxelocks_simple_sku"]
            if m["velourahair_simple_sku"]:
                map_veloura[ss] = m["velourahair_simple_sku"]

        # ── 组装结果，应用映射 ──
        result = []
        for p in products:
            pd = dict(p)
            ss = simplify_sku(pd.get("sku", ""))
            pd["simple_sku"] = ss

            # 跨境: 用映射后的SKU，无映射则用原SKU
            cross_key = map_cross.get(ss, ss)
            pd["cross_available"] = wh_cross.get(cross_key, {}).get("available", 0)
            pd["cross_transit"] = wh_cross.get(cross_key, {}).get("transit", 0)

            luxe_key = map_luxe.get(ss, ss)
            pd["luxe_available"] = wh_luxe.get(luxe_key, {}).get("available", 0)
            pd["luxe_transit"] = wh_luxe.get(luxe_key, {}).get("transit", 0)

            veloura_key = map_veloura.get(ss, ss)
            pd["veloura_available"] = wh_veloura.get(veloura_key, {}).get("available", 0)
            pd["veloura_transit"] = wh_veloura.get(veloura_key, {}).get("transit", 0)

            result.append(pd)

        # Python层面排序（仓库列是计算出来的，SQL无法排）
        sort_keys = {
            "shopify": "inventory_quantity",
            "cross_avail": "cross_available",
            "cross_transit": "cross_transit",
            "luxe_avail": "luxe_available",
            "luxe_transit": "luxe_transit",
            "veloura_avail": "veloura_available",
            "veloura_transit": "veloura_transit",
        }
        if sort and sort in sort_keys:
            key = sort_keys[sort]
            reverse = order != "asc"
            result.sort(key=lambda p: p.get(key, 0) or 0, reverse=reverse)

        # 计算总计
        totals = {
            "shopify": sum(p["inventory_quantity"] or 0 for p in result),
            "cross_available": sum(p["cross_available"] or 0 for p in result),
            "cross_transit": sum(p["cross_transit"] or 0 for p in result),
            "luxe_available": sum(p["luxe_available"] or 0 for p in result),
            "luxe_transit": sum(p["luxe_transit"] or 0 for p in result),
            "veloura_available": sum(p["veloura_available"] or 0 for p in result),
            "veloura_transit": sum(p["veloura_transit"] or 0 for p in result),
        }

        resp = render_html("products.html", request, products=result, totals=totals, search=search, sort=sort, order=order, status=status)
        resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        return resp
    finally:
        conn.close()


# ────────────────────────────────────────
# 头程管理
# ────────────────────────────────────────
@app.get("/headhaul", response_class=HTMLResponse)
async def headhaul_page(request: Request):
    conn = get_db()
    try:
        orders = conn.execute("SELECT * FROM headhaul_orders ORDER BY created_at DESC LIMIT 100").fetchall()
        total = conn.execute("SELECT COUNT(*) FROM headhaul_orders").fetchone()[0]
        in_transit = conn.execute("SELECT COUNT(*) FROM headhaul_orders WHERE status='in_transit'").fetchone()[0]
        delivered = conn.execute("SELECT COUNT(*) FROM headhaul_orders WHERE status='delivered'").fetchone()[0]
        return render_html("headhaul.html", request,
            orders=[dict(o) for o in orders],
            stats={"total": total, "in_transit": in_transit, "delivered": delivered}
        )
    finally:
        conn.close()


# ────────────────────────────────────────
# 领星 OMS 连接器
# ────────────────────────────────────────
OMS_BASE = "https://api.xlwms.com/openapi"
OMS_AK = OMS_APP_KEY
OMS_SK = OMS_APP_SECRET

def oms_sign(app_key: str, app_secret: str, data: dict, req_time: str) -> str:
    """领星 OMS HMAC-SHA256 签名"""
    # Step 1: sort data fields
    sorted_data = dict(sorted(data.items(), key=lambda x: x[0].lower()))
    import json
    data_json = json.dumps(sorted_data, separators=(',', ':'), ensure_ascii=False)
    # Step 2: sort top-level params and concat
    parts = {"appKey": app_key, "data": data_json, "reqTime": req_time}
    sign_str = "".join(v for k, v in sorted(parts.items(), key=lambda x: x[0].lower()))
    # Step 3: HMAC-SHA256
    import hmac as hmac_mod
    sig = hmac_mod.new(app_secret.encode(), sign_str.encode(), hashlib.sha256).hexdigest()
    return sig

async def oms_call(endpoint: str, data: dict) -> dict:
    """调用领星 OMS API"""
    import time as time_mod
    req_time = str(int(time_mod.time()))
    sign = oms_sign(OMS_AK, OMS_SK, data, req_time)
    url = f"{OMS_BASE}{endpoint}?authcode={sign}"
    body = {"appKey": OMS_AK, "reqTime": req_time, "data": data}
    async with httpx.AsyncClient() as client:
        resp = await client.post(url, json=body, timeout=20)
        return resp.json()


@app.post("/api/oms/sync")
async def sync_oms_tracking():
    """从 OMS 同步物流单号，匹配 Hub 订单（分批拉取）"""
    conn = get_db()
    try:
        total_synced = 0
        total_matched = 0
        total_updated = 0

        result = await oms_call("/v1/outboundOrder/pageList", {"page": 1, "pageSize": 20})
        if result.get("code") != 200:
            return {"status": "error", "message": f"OMS返回: {result}"}

        records = result.get("data", {}).get("records", [])
        total_synced = len(records)

        for rec in records:
                track_no = rec.get("logisticsTrackNo", "")
                platform_order_no = rec.get("platformOrderNo", "")
                oms_status = rec.get("status", 0)

                if not platform_order_no:
                    continue

                local = conn.execute(
                    "SELECT * FROM orders WHERE platform_order_id = ? OR order_number = ?",
                    (str(platform_order_no), str(platform_order_no))
                ).fetchone()

                if local:
                    total_matched += 1
                    if track_no:
                        new_status = "shipped" if oms_status == 3 else "paid"
                        conn.execute("""
                            UPDATE orders SET tracking_number = ?, tracking_company = ?,
                            status = ?, updated_at = CURRENT_TIMESTAMP
                            WHERE platform_order_id = ? OR order_number = ?
                        """, (track_no, rec.get("logisticsChannel", "OMS"),
                              new_status, str(platform_order_no), str(platform_order_no)))
                        total_updated += 1

        conn.commit()
        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("oms", "sync_tracking", "success",
             f"Synced {total_synced} records, matched {total_matched}, updated {total_updated}")
        )
        conn.commit()
        return {
            "status": "ok",
            "total_oms": total_synced,
            "matched": total_matched,
            "updated": total_updated
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/api/oms/sync-tracking")
async def sync_oms_tracking_overseas():
    """从OMS同步运单号 — 仅海外仓发货订单，按订单号匹配"""
    conn = get_db()
    try:
        # 收集所有海外仓发货的订单号（不限制是否已有运单号）
        overseas = conn.execute(
            "SELECT id, order_number, tracking_number FROM orders"
            " WHERE shipping_type = '海外仓发货'"
        ).fetchall()
        if not overseas:
            return {"status": "ok", "message": "没有海外仓订单", "updated": 0}

        order_map = {o["order_number"]: o for o in overseas if o["order_number"]}
        if not order_map:
            return {"status": "ok", "message": "没有有效的订单号", "updated": 0}

        # 拉取最近30天OMS出库单
        from datetime import timedelta
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        start_str = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d %H:%M:%S")

        updated = 0
        matched_count = 0
        skipped = 0
        page = 1
        while page <= 5:
            result = await oms_call("/v1/outboundOrder/pageList", {
                "page": page,
                "pageSize": 50,
                "timeType": "orderCreateTime",
                "startTime": start_str,
                "endTime": now_str
            })
            if result.get("code") != 200:
                break

            records = result.get("data", {}).get("records", [])
            if not records:
                break

            for rec in records:
                platform_no = rec.get("platformOrderNo", "")
                track_no = rec.get("logisticsTrackNo", "")
                channel = rec.get("logisticsChannel", "OMS")
                oms_status = rec.get("status", 0)

                if not platform_no or not track_no:
                    continue

                if platform_no in order_map:
                    matched_count += 1
                    existing = order_map[platform_no]
                    # 如果本地已有运单号且不同，跳过（保留手动录入的）
                    if existing["tracking_number"] and existing["tracking_number"] != track_no:
                        skipped += 1
                        continue

                    new_status = "shipped" if oms_status == 3 else None
                    if new_status:
                        conn.execute(
                            "UPDATE orders SET tracking_number = ?, tracking_company = ?,"
                            " status = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                            (track_no, channel, new_status, existing["id"])
                        )
                    else:
                        conn.execute(
                            "UPDATE orders SET tracking_number = ?, tracking_company = ?,"
                            " updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                            (track_no, channel, existing["id"])
                        )
                    updated += 1
                    del order_map[platform_no]

            page += 1

        conn.commit()
        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("oms", "sync_tracking_overseas", "success",
             f"Matched {matched_count}, updated {updated}, skipped {skipped}")
        )
        conn.commit()
        return {
            "status": "ok",
            "matched": matched_count,
            "updated": updated,
            "skipped": skipped,
            "message": f"同步完成: 匹配 {matched_count} 单, 更新 {updated} 单运单号" +
                       (f", 跳过 {skipped} 单(已有运单)" if skipped else "")
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/api/oms/sync-inventory")
async def sync_oms_inventory():
    """从 OMS 同步跨境库存（可用+在途）到产品表"""
    conn = get_db()
    try:
        page = 1
        total_synced = 0
        while page <= 10:
            result = await oms_call("/v1/integratedInventory/pageOpen", {"page": page, "pageSize": 50})
            if result.get("code") != 200:
                break
            records = result.get("data", {}).get("records", [])
            if not records:
                break

            for rec in records:
                oms_sku = rec.get("sku", "")
                if not oms_sku:
                    continue
                # 用简化SKU匹配
                simple_sku = simplify_sku(oms_sku)
                available = int(rec.get("productStockDtl", {}).get("availableAmount", 0) or 0)
                transit = int(rec.get("productStockDtl", {}).get("transportAmount", 0) or 0)

                # 匹配本地产品（简化SKU 或 原始SKU）
                local = conn.execute(
                    "SELECT sku FROM products WHERE sku = ?", (oms_sku,)
                ).fetchone()
                if local:
                    conn.execute("""
                        UPDATE products SET oms_available_qty = ?, oms_transit_qty = ?,
                        updated_at = CURRENT_TIMESTAMP WHERE sku = ?
                    """, (available, transit, oms_sku))
                else:
                    # 用简化SKU匹配本地产品
                    conn.execute("""
                        UPDATE products SET oms_available_qty = ?, oms_transit_qty = ?,
                        updated_at = CURRENT_TIMESTAMP WHERE sku LIKE ? OR sku = ?
                    """, (available, transit, f"%{simple_sku}%", oms_sku))
                total_synced += 1

            total_pages = result.get("data", {}).get("pages", 1)
            if page >= total_pages:
                break
            page += 1

        conn.commit()
        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("oms", "sync_inventory", "success", f"Synced {total_synced} SKUs")
        )
        conn.commit()
        return {"status": "ok", "synced": total_synced}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/api/inventory/sync-cross-border")
async def sync_cross_border_from_oms():
    """从领星OMS同步库存到跨境在线仓(warehouse_inventory)"""
    conn = get_db()
    try:
        all_records = []
        page = 1
        while page <= 10:
            result = await oms_call("/v1/integratedInventory/pageOpen", {"page": page, "pageSize": 50})
            if result.get("code") != 200:
                break
            records = result.get("data", {}).get("records", [])
            if not records:
                break
            all_records.extend(records)
            total_pages = result.get("data", {}).get("pages", 1)
            if page >= total_pages:
                break
            page += 1

        # 按SKU汇总
        aggregated = {}
        for rec in all_records:
            sku = rec.get("sku", "").strip()
            if not sku:
                continue
            available = int(rec.get("productStockDtl", {}).get("availableAmount", 0) or 0)
            transit = int(rec.get("productStockDtl", {}).get("transportAmount", 0) or 0)
            if sku in aggregated:
                aggregated[sku]["available"] += available
                aggregated[sku]["transit"] += transit
            else:
                aggregated[sku] = {"reference_code": sku, "available": available, "transit": transit}

        # UPSERT: OMS有的更新，OMS没有的保留（手动添加的新品不被覆盖）
        updated = 0
        inserted = 0
        for sku, data in aggregated.items():
            existing = conn.execute(
                "SELECT id FROM warehouse_inventory WHERE warehouse_name='CrossBorder' AND reference_code=?",
                (data["reference_code"],)
            ).fetchone()
            if existing:
                conn.execute("""
                    UPDATE warehouse_inventory SET available_inventory=?, in_transit_total=?, updated_at=CURRENT_TIMESTAMP
                    WHERE id=?
                """, (data["available"], data["transit"], existing["id"]))
                updated += 1
            else:
                conn.execute("""
                    INSERT INTO warehouse_inventory (reference_code, warehouse_name, available_inventory, in_transit_total)
                    VALUES (?, 'CrossBorder', ?, ?)
                """, (data["reference_code"], data["available"], data["transit"]))
                inserted += 1
        conn.commit()

        # 统计手动SKU数量
        manual_count = conn.execute(
            "SELECT COUNT(*) FROM warehouse_inventory WHERE warehouse_name='CrossBorder'"
        ).fetchone()[0] - len(aggregated)

        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("oms", "sync_cross_border", "success",
             f"OMS: {len(all_records)} rows → {len(aggregated)} SKUs (update {updated}, insert {inserted}), 手动保留 {max(0, manual_count)}")
        )
        conn.commit()

        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("oms", "sync_cross_border", "success",
             f"OMS: {len(all_records)} rows -> {len(aggregated)} SKUs (update {updated}, insert {inserted}),"
             f" manual {max(0, manual_count)}")
        )
        conn.commit()
        return {"status": "ok", "synced": len(aggregated), "total_rows": len(all_records),
                "updated": updated, "inserted": inserted, "manual_kept": max(0, manual_count)}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/api/inventory/push-to-shopify")
async def push_inventory_to_shopify():
    """将跨境在线库存可用同步到 Shopify（仅更新有库存的SKU，不回写0）"""
    conn = get_db()
    try:
        global _shopify_location_id
        if not _shopify_location_id:
            loc_result = await shopify.get("/locations.json")
            locs = loc_result.get("locations", [])
            if locs:
                _shopify_location_id = locs[0]["id"]
        if not _shopify_location_id:
            return {"status": "error", "message": "无法获取Shopify仓库位置"}

        # SKU映射
        mappings = conn.execute(
            "SELECT shopify_simple_sku, cross_border_simple_sku FROM sku_mapping"
        ).fetchall()
        map_cross = {}
        for m in mappings:
            if m["cross_border_simple_sku"]:
                map_cross[m["shopify_simple_sku"]] = m["cross_border_simple_sku"]

        # 查询跨境在线库存（按简化SKU汇总，含avail=0）
        cross_inv = {}  # simplified_sku → available qty
        for row in conn.execute(
            "SELECT reference_code, SUM(available_inventory) as total FROM warehouse_inventory"
            " WHERE warehouse_name IN ('CrossBorder','NewProducts')"
            " GROUP BY reference_code"
        ).fetchall():
            ref_simple = simplify_sku(row["reference_code"])
            if ref_simple:
                cross_inv[ref_simple] = cross_inv.get(ref_simple, 0) + (row["total"] or 0)

        # 收集所有OMS管理的简化SKU（有warehouse_inventory记录的，无论库存多少）
        oms_managed_skus = set(cross_inv.keys())

        # 查询所有有variant_id的产品
        products = conn.execute(
            "SELECT sku, inventory_item_id, variant_id FROM products"
            " WHERE inventory_item_id IS NOT NULL AND inventory_item_id != ''"
        ).fetchall()

        updated = 0
        untracked = 0
        skipped = 0
        errors = 0
        error_skus = []
        for p in products:
            inv_id = p["inventory_item_id"]
            if not inv_id:
                continue
            ss = simplify_sku(p["sku"])
            cross_key = map_cross.get(ss, ss)

            # 只处理OMS管理的SKU
            if cross_key not in oms_managed_skus:
                skipped += 1
                continue

            avail = cross_inv.get(cross_key, 0)
            vid = int(p["variant_id"]) if p["variant_id"] and p["variant_id"].isdigit() else None

            try:
                if avail > 0:
                    # 有库存 → 跟踪 + 设置库存数
                    await shopify.set_inventory(int(inv_id), _shopify_location_id, int(avail), vid)
                    conn.execute("UPDATE products SET inventory_quantity=? WHERE sku=?", (int(avail), p["sku"]))
                    updated += 1
                else:
                    # 库存耗尽 → 取消跟踪，继续卖（未跟踪库存）
                    if vid:
                        await shopify.untrack_variant(vid)
                    untracked += 1
                    conn.execute("UPDATE products SET inventory_quantity=0 WHERE sku=?", (p["sku"],))

                if (updated + untracked) % 10 == 0:
                    await asyncio.sleep(0.5)
            except Exception as e:
                errors += 1
                error_skus.append(f"{ss}({str(e)[:50]})")

        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("oms", "push_shopify_inventory", "success",
             f"Shopify push: {updated} updated, {untracked} untracked, {skipped} skipped, {errors} errors")
        )
        conn.commit()
        err_detail = ", 失败: " + "; ".join(error_skus) if error_skus else ""
        return {"status": "ok", "updated": updated, "untracked": untracked,
                "skipped": skipped, "errors": errors,
                "message": f"Shopify同步: 跟踪更新 {updated} SKU, 转为未跟踪 {untracked} SKU (库存耗尽继续卖), 跳过 {skipped}, 失败 {errors}" + err_detail}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.get("/api/oms/stats")
async def oms_stats():
    """OMS 出库统计"""
    result = await oms_call("/v1/outboundOrder/pageList", {
        "page": 1, "pageSize": 1,
        "timeType": "orderCreateTime",
        "startTime": (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d %H:%M:%S"),
        "endTime": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })
    total = result.get("data", {}).get("total", 0) if result.get("code") == 200 else 0
    return {"total_oms_orders": total}


# ────────────────────────────────────────
# 采购资源信息池
# ────────────────────────────────────────
@app.get("/procurement", response_class=HTMLResponse)
async def procurement_page(request: Request, search: str = None):
    """采购资源信息池页面"""
    conn = get_db()
    try:
        query = "SELECT * FROM procurement_resources WHERE 1=1"
        params = []
        if search:
            query += " AND (simple_sku LIKE ? OR supplier LIKE ? OR contact_name LIKE ? OR notes LIKE ? OR seller_id LIKE ?)"
            st = f"%{search}%"
            params.extend([st, st, st, st, st])
        query += " ORDER BY updated_at DESC LIMIT 200"
        rows = conn.execute(query, params).fetchall()

        # 加载供应商数据：seller_id → {purchase_link, contact_name, wechat}
        supplier_rows = conn.execute(
            "SELECT name, seller_id, purchase_link, contact_name, wechat FROM suppliers"
        ).fetchall()
        sup_map = {}
        for s in supplier_rows:
            sid = (s["seller_id"] or "").strip()
            if sid:
                sup_map[sid] = {
                    "supplier_name": s["name"],
                    "purchase_link": s["purchase_link"] or "",
                    "contact_name": s["contact_name"] or "",
                    "wechat": s["wechat"] or ""
                }

        # 为每个资源合并供应商数据
        resources = []
        for r in rows:
            rd = dict(r)
            sid = (rd.get("seller_id") or "").strip()
            sup = sup_map.get(sid)
            if sup:
                rd["supplier_name"] = sup["supplier_name"]
                # 供应商数据覆盖（如果资源自身没有值）
                if not rd.get("purchase_link"):
                    rd["purchase_link"] = sup["purchase_link"]
                if not rd.get("contact_name"):
                    rd["contact_name"] = sup["contact_name"]
                if not rd.get("wechat"):
                    rd["wechat"] = sup["wechat"]
            resources.append(rd)

        # 供应商列表（供新增/编辑下拉选择）
        all_suppliers = conn.execute(
            "SELECT id, name, seller_id FROM suppliers ORDER BY name"
        ).fetchall()

        return render_html("procurement.html", request,
            resources=resources,
            suppliers=[dict(s) for s in all_suppliers],
            search=search
        )
    finally:
        conn.close()


@app.get("/api/procurement/list")
async def procurement_list(search: str = ""):
    """API: 采购资源列表"""
    conn = get_db()
    try:
        query = "SELECT * FROM procurement_resources WHERE 1=1"
        params = []
        if search:
            query += " AND (simple_sku LIKE ? OR supplier LIKE ? OR contact_name LIKE ?)"
            st = f"%{search}%"
            params.extend([st, st, st])
        query += " ORDER BY updated_at DESC LIMIT 200"
        rows = conn.execute(query, params).fetchall()
        return {"resources": [dict(r) for r in rows]}
    finally:
        conn.close()


@app.post("/api/procurement/add")
async def procurement_add(request: Request):
    """添加采购资源（重复SKU自动变为更新）"""
    data = await request.json()
    conn = get_db()
    try:
        sku = data.get("simple_sku", "").strip()
        if not sku:
            return {"status": "error", "message": "简化SKU必填"}
        # 检查是否已存在
        existing = conn.execute(
            "SELECT id FROM procurement_resources WHERE simple_sku = ?", (sku,)
        ).fetchone()
        if existing:
            # 已存在 → 更新
            conn.execute("""
                UPDATE procurement_resources SET
                    image_url = ?, supplier = ?, cost = ?,
                    estimated_delivery = ?, purchase_link = ?,
                    contact_name = ?, wechat = ?, notes = ?,
                    seller_id = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                data.get("image_url", "").strip(),
                data.get("supplier", "").strip(),
                float(data.get("cost", 0) or 0),
                data.get("estimated_delivery", "").strip(),
                data.get("purchase_link", "").strip(),
                data.get("contact_name", "").strip(),
                data.get("wechat", "").strip(),
                data.get("notes", "").strip(),
                data.get("seller_id", "").strip(),
                existing["id"]
            ))
            conn.commit()
            return {"status": "ok", "message": f"已更新 {sku}（该SKU已存在）"}
        else:
            conn.execute("""
                INSERT INTO procurement_resources
                (simple_sku, image_url, supplier, cost, estimated_delivery,
                 purchase_link, contact_name, wechat, notes, seller_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                sku,
                data.get("image_url", "").strip(),
                data.get("supplier", "").strip(),
                float(data.get("cost", 0) or 0),
                data.get("estimated_delivery", "").strip(),
                data.get("purchase_link", "").strip(),
                data.get("contact_name", "").strip(),
                data.get("wechat", "").strip(),
                data.get("notes", "").strip(),
                data.get("seller_id", "").strip()
            ))
            conn.commit()
            return {"status": "ok", "message": "已添加"}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/api/procurement/update")
async def procurement_update(request: Request):
    """更新采购资源"""
    data = await request.json()
    conn = get_db()
    try:
        rid = data.get("id")
        if not rid:
            return {"status": "error", "message": "缺少ID"}
        conn.execute("""
            UPDATE procurement_resources SET
                simple_sku = ?, image_url = ?, supplier = ?, cost = ?,
                estimated_delivery = ?, purchase_link = ?, contact_name = ?,
                wechat = ?, notes = ?, seller_id = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (
            data.get("simple_sku", "").strip(),
            data.get("image_url", "").strip(),
            data.get("supplier", "").strip(),
            float(data.get("cost", 0) or 0),
            data.get("estimated_delivery", "").strip(),
            data.get("purchase_link", "").strip(),
            data.get("contact_name", "").strip(),
            data.get("wechat", "").strip(),
            data.get("notes", "").strip(),
            data.get("seller_id", "").strip(),
            rid
        ))
        conn.commit()
        return {"status": "ok", "message": "已更新"}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/api/procurement/delete")
async def procurement_delete(request: Request):
    """删除采购资源"""
    data = await request.json()
    conn = get_db()
    try:
        conn.execute("DELETE FROM procurement_resources WHERE id = ?", (data.get("id"),))
        conn.commit()
        return {"status": "ok", "message": "已删除"}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.get("/api/procurement/import-from-product")
async def procurement_import_product(sku: str = ""):
    """从产品库按简化SKU导入信息"""
    conn = get_db()
    try:
        if not sku:
            return {"status": "error", "message": "请输入简化SKU"}
        # 匹配产品: SKU包含该简化SKU
        product = conn.execute(
            "SELECT sku, title, image_url FROM products WHERE sku LIKE ? LIMIT 1",
            (f"%{sku}%",)
        ).fetchone()
        if not product:
            return {"status": "error", "message": f"未找到SKU包含 '{sku}' 的产品"}
        return {
            "status": "ok",
            "simple_sku": simplify_sku(product["sku"]),
            "image_url": product["image_url"] or "",
            "product_title": product["title"] or "",
            "product_sku": product["sku"]
        }
    finally:
        conn.close()


@app.get("/api/procurement/product-list")
async def procurement_product_list(search: str = ""):
    """产品列表供采购池勾选导入（标记已导入的SKU，防止重复）"""
    conn = get_db()
    try:
        # 查询产品
        query = "SELECT sku, title, image_url FROM products WHERE 1=1"
        params = []
        if search:
            query += " AND (sku LIKE ? OR title LIKE ?)"
            st = f"%{search}%"
            params.extend([st, st])
        query += " ORDER BY title LIMIT 200"
        rows = conn.execute(query, params).fetchall()

        # 查询已导入的简化SKU集合
        existing_rows = conn.execute("SELECT simple_sku FROM procurement_resources").fetchall()
        existing_set = {r["simple_sku"] for r in existing_rows}

        products = []
        for p in rows:
            ss = simplify_sku(p["sku"])
            if not ss:
                continue
            already_imported = ss in existing_set
            products.append({
                "simple_sku": ss,
                "full_sku": p["sku"],
                "title": p["title"] or "",
                "image_url": p["image_url"] or "",
                "already_imported": already_imported
            })
        return {"status": "ok", "products": products}
    finally:
        conn.close()


# ────────────────────────────────────────
# TikTok Wig Ops 中控
# ────────────────────────────────────────
# TikTok SKU管理
# ────────────────────────────────────────
@app.get("/tiktok-sku", response_class=HTMLResponse)
async def tiktok_sku_page(request: Request):
    """TikTok SKU映射管理"""
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM tiktok_sku_mapping ORDER BY COALESCE(simple_sku, sku), sku"
        ).fetchall()
        return render_html("tiktok_sku.html", request, mappings=[dict(r) for r in rows])
    finally:
        conn.close()


@app.post("/api/tiktok-sku/save")
async def tiktok_sku_mapping_save(request: Request):
    data = await request.json()
    product_id = str(data.get("tiktok_product_id", "")).strip()
    sku = str(data.get("sku", "")).strip()
    if not product_id or not sku:
        return {"status": "error", "message": "TikTok商品ID和SKU必填"}
    simple = str(data.get("simple_sku", "")).strip() or simplify_sku(sku)
    conn = get_db()
    try:
        values = (
            product_id, sku, simple,
            safe_float(data.get("price")), safe_float(data.get("product_cost")),
            safe_float(data.get("shipping_fee")), safe_float(data.get("platform_fee_rate"), 0.06),
            safe_float(data.get("ad_cost")), safe_float(data.get("refund_loss")),
            safe_float(data.get("return_rate"), 0.20), str(data.get("notes", "")).strip()
        )
        if data.get("id"):
            conn.execute("""
                UPDATE tiktok_sku_mapping SET
                    tiktok_product_id=?, sku=?, simple_sku=?, price=?, product_cost=?,
                    shipping_fee=?, platform_fee_rate=?, ad_cost=?, refund_loss=?,
                    return_rate=?, notes=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            """, values + (safe_int(data.get("id")),))
        else:
            conn.execute("""
                INSERT INTO tiktok_sku_mapping
                    (tiktok_product_id, sku, simple_sku, price, product_cost,
                     shipping_fee, platform_fee_rate, ad_cost, refund_loss,
                     return_rate, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(tiktok_product_id) DO UPDATE SET
                    sku=excluded.sku,
                    simple_sku=excluded.simple_sku,
                    price=excluded.price,
                    product_cost=excluded.product_cost,
                    shipping_fee=excluded.shipping_fee,
                    platform_fee_rate=excluded.platform_fee_rate,
                    ad_cost=excluded.ad_cost,
                    refund_loss=excluded.refund_loss,
                    return_rate=excluded.return_rate,
                    notes=excluded.notes,
                    updated_at=CURRENT_TIMESTAMP
            """, values)
        ensure_tiktok_ops_sku(conn, {
            "sku": sku,
            "simple_sku": simple,
            "price": data.get("price"),
            "product_cost": data.get("product_cost"),
            "shipping_fee": data.get("shipping_fee"),
            "platform_fee_rate": data.get("platform_fee_rate"),
            "ad_cost": data.get("ad_cost"),
            "refund_loss": data.get("refund_loss"),
            "return_rate": data.get("return_rate"),
            "notes": data.get("notes")
        })
        conn.commit()
        return {"status": "ok"}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/api/tiktok-sku/delete")
async def tiktok_sku_mapping_delete(request: Request):
    data = await request.json()
    rid = safe_int(data.get("id"))
    conn = get_db()
    try:
        conn.execute("DELETE FROM tiktok_sku_mapping WHERE id=?", (rid,))
        conn.commit()
        return {"status": "ok"}
    finally:
        conn.close()


# 列表需要的列（避免 SELECT * 加载40列）
_VIDEO_LIST_COLS = [
    "id","creator_nickname","video_info","video_id","publish_time",
    "product_name","tiktok_product_id","sku","simple_sku",
    "vv","product_clicks","attributed_sku_orders","attributed_gmv",
    "platform_diagnosis","local_diagnosis","repeat_action"
]
_VIDEO_LIST_SELECT = ",".join(_VIDEO_LIST_COLS)

@app.get("/tiktok-videos", response_class=HTMLResponse)
async def tiktok_videos_page(request: Request, search: str = "", mapped: str = "all", page: int = 1):
    """TikTok 视频表现导入表（分页，每页50条）"""
    conn = get_db()
    try:
        per_page = 50
        # 允许排序的白名单
        allowed_sort = {"publish_time", "vv", "product_clicks", "attributed_sku_orders", "attributed_gmv", "id"}
        sort = request.query_params.get("sort", "publish_time")
        order = request.query_params.get("order", "desc")
        if sort not in allowed_sort:
            sort = "publish_time"
        if order not in ("asc", "desc"):
            order = "desc"

        where = ["1=1"]
        params = []
        if search:
            st = f"%{search}%"
            where.append(
                "(creator_nickname LIKE ? OR video_info LIKE ? OR video_id LIKE ? OR "
                "product_name LIKE ? OR tiktok_product_id LIKE ? OR sku LIKE ? OR simple_sku LIKE ?)"
            )
            params.extend([st, st, st, st, st, st, st])
        if mapped == "mapped":
            where.append("sku NOT LIKE 'TT-%'")
        elif mapped == "unmapped":
            where.append("sku LIKE 'TT-%'")
        where_clause = " AND ".join(where)

        # 总数
        total = conn.execute(
            f"SELECT COUNT(*) FROM tiktok_video_performance WHERE {where_clause}", params
        ).fetchone()[0]
        total_pages = max(1, (total + per_page - 1) // per_page)
        offset = max(0, min(page, total_pages) - 1) * per_page

        rows = conn.execute(
            f"SELECT {_VIDEO_LIST_SELECT} FROM tiktok_video_performance WHERE {where_clause} "
            f"ORDER BY {sort} {order} LIMIT ? OFFSET ?",
            params + [per_page, offset]
        ).fetchall()

        stats = conn.execute("""
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN sku LIKE 'TT-%' THEN 1 ELSE 0 END) as unmapped,
                SUM(vv) as views,
                SUM(product_clicks) as clicks,
                SUM(attributed_sku_orders) as orders,
                SUM(attributed_gmv) as gmv,
                MAX(imported_at) as latest_import
            FROM tiktok_video_performance
        """).fetchone()
        stats_d = dict(stats)
        stats_d["mapped"] = (stats_d.get("total") or 0) - (stats_d.get("unmapped") or 0)
        stats_d["ctr"] = round((stats_d.get("clicks") or 0) / max(stats_d.get("views") or 0, 1) * 100, 2)
        return render_html(
            "tiktok_videos.html",
            request,
            rows=[dict(r) for r in rows],
            stats=stats_d,
            search=search,
            mapped=mapped,
            page=page,
            total_pages=total_pages,
            total=total,
            sort=sort,
            order=order
        )
    finally:
        conn.close()


@app.post("/api/tiktok-videos/import")
async def tiktok_videos_import(file: UploadFile = File(...)):
    """导入 TikTok Video Performance List xlsx，并同步到 Wig Ops 复盘表"""
    filename = file.filename or "tiktok_video_performance.xlsx"
    if not filename.lower().endswith(".xlsx"):
        return {"status": "error", "message": "请上传 .xlsx 文件"}
    content = await file.read()
    conn = get_db()
    try:
        headers, raw_rows = read_xlsx_first_sheet(content)
        required = ["达人昵称", "达人ID", "视频信息", "视频ID", "发布时间", "商品", "VV", "商品点击次数"]
        missing = [h for h in required if h not in headers]
        if missing:
            return {"status": "error", "message": "缺少字段: " + ", ".join(missing)}
        map_rows = conn.execute("SELECT * FROM tiktok_sku_mapping").fetchall()
        mapping = {str(r["tiktok_product_id"]): dict(r) for r in map_rows}
        imported = 0
        synced = 0
        unmapped = 0
        skipped = 0
        for raw in raw_rows:
            item = tiktok_perf_from_export_row(raw, mapping, filename)
            if not item["video_id"]:
                skipped += 1
                continue
            upsert_tiktok_video_performance(conn, item)
            imported += 1
            if item["sku"].startswith("TT-"):
                unmapped += 1
            sync_perf_to_tiktok_videos(conn, item)
            synced += 1
        conn.commit()
        return {
            "status": "ok",
            "headers": headers,
            "imported": imported,
            "synced": synced,
            "unmapped": unmapped,
            "skipped": skipped,
            "message": f"导入 {imported} 行，同步复盘 {synced} 行，未映射 {unmapped} 行"
        }
    except Exception as e:
        conn.rollback()
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


# ────────────────────────────────────────
@app.get("/tiktok", response_class=HTMLResponse)
async def tiktok_ops_page(request: Request):
    """TikTok 假发运营中控页面"""
    conn = get_db()
    try:
        sku_rows = get_tiktok_sku_options(conn)
        skus = []
        high_risk = 0
        main_push = 0
        low_stock = 0
        for row in sku_rows:
            item = dict(row)
            metrics = calc_tiktok_profit(item)
            item.update(metrics)
            if metrics["profit_label"] == "高风险":
                high_risk += 1
            if metrics["profit_label"] == "可主推":
                main_push += 1
            if metrics["stock_status"] in ("需补货", "低库存"):
                low_stock += 1
            skus.append(item)

        video_rows = conn.execute(
            "SELECT * FROM tiktok_videos ORDER BY publish_date DESC, id DESC LIMIT 200"
        ).fetchall()
        videos = []
        repeat_candidates = 0
        for row in video_rows:
            item = dict(row)
            metrics = diagnose_tiktok_video(item)
            if not item.get("diagnosis"):
                item["diagnosis"] = metrics["diagnosis"]
            if not item.get("repeat_action"):
                item["repeat_action"] = metrics["repeat_action"]
            item.update({k: v for k, v in metrics.items() if k not in ("diagnosis", "repeat_action")})
            if item["diagnosis"] in ("可复拍模板", "小流量高转化"):
                repeat_candidates += 1
            videos.append(item)

        today = datetime.now().strftime("%Y-%m-%d")
        upcoming = conn.execute(
            "SELECT COUNT(*) FROM tiktok_videos WHERE posted=0 AND publish_date >= ?",
            (today,)
        ).fetchone()[0]

        stats = {
            "total_skus": len(skus),
            "main_push": main_push,
            "high_risk": high_risk,
            "low_stock": low_stock,
            "upcoming": upcoming,
            "repeat_candidates": repeat_candidates
        }
        return render_html(
            "tiktok.html",
            request,
            skus=skus,
            videos=videos,
            stats=stats,
            today=today,
            default_accounts=", ".join(TIKTOK_DEFAULT_ACCOUNTS),
            angle_pool=TIKTOK_ANGLE_POOL
        )
    finally:
        conn.close()


@app.get("/api/tiktok/sku-options")
async def tiktok_sku_options(search: str = ""):
    conn = get_db()
    try:
        options = get_tiktok_sku_options(conn)
        if search:
            s = search.lower()
            options = [o for o in options if s in str(o.get("search_text") or "").lower()]
        return {"status": "ok", "options": options[:200]}
    finally:
        conn.close()


@app.post("/api/tiktok/sku/save")
async def tiktok_sku_save(request: Request):
    data = await request.json()
    sku = data.get("sku", "").strip()
    if not sku:
        return {"status": "error", "message": "SKU必填"}
    conn = get_db()
    try:
        values = (
            sku,
            data.get("product_name", "").strip(),
            data.get("color", "").strip(),
            data.get("length", "").strip(),
            safe_float(data.get("price"), 23.99),
            safe_float(data.get("product_cost")),
            safe_float(data.get("shipping_fee")),
            safe_float(data.get("platform_fee_rate"), 0.06),
            safe_float(data.get("ad_cost")),
            safe_float(data.get("return_rate"), 0.20),
            safe_float(data.get("refund_loss")),
            safe_int(data.get("stock")),
            safe_float(data.get("daily_sales")),
            safe_int(data.get("lead_time_days"), 30),
            safe_int(data.get("safety_stock"), 10),
            data.get("status", "testing").strip() or "testing",
            data.get("notes", "").strip()
        )
        if data.get("id"):
            conn.execute("""
                UPDATE tiktok_skus SET
                    sku=?, product_name=?, color=?, length=?, price=?,
                    product_cost=?, shipping_fee=?, platform_fee_rate=?, ad_cost=?,
                    return_rate=?, refund_loss=?, stock=?, daily_sales=?,
                    lead_time_days=?, safety_stock=?, status=?, notes=?,
                    updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            """, values + (safe_int(data.get("id")),))
        else:
            conn.execute("""
                INSERT INTO tiktok_skus
                    (sku, product_name, color, length, price,
                     product_cost, shipping_fee, platform_fee_rate, ad_cost,
                     return_rate, refund_loss, stock, daily_sales,
                     lead_time_days, safety_stock, status, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(sku) DO UPDATE SET
                    product_name=excluded.product_name,
                    color=excluded.color,
                    length=excluded.length,
                    price=excluded.price,
                    product_cost=excluded.product_cost,
                    shipping_fee=excluded.shipping_fee,
                    platform_fee_rate=excluded.platform_fee_rate,
                    ad_cost=excluded.ad_cost,
                    return_rate=excluded.return_rate,
                    refund_loss=excluded.refund_loss,
                    stock=excluded.stock,
                    daily_sales=excluded.daily_sales,
                    lead_time_days=excluded.lead_time_days,
                    safety_stock=excluded.safety_stock,
                    status=excluded.status,
                    notes=excluded.notes,
                    updated_at=CURRENT_TIMESTAMP
            """, values)
        conn.commit()
        return {"status": "ok"}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/api/tiktok/sku/delete")
async def tiktok_sku_delete(request: Request):
    data = await request.json()
    rid = safe_int(data.get("id"))
    conn = get_db()
    try:
        conn.execute("DELETE FROM tiktok_skus WHERE id=?", (rid,))
        conn.commit()
        return {"status": "ok"}
    finally:
        conn.close()


@app.post("/api/tiktok/video/save")
async def tiktok_video_save(request: Request):
    data = await request.json()
    sku = data.get("sku", "").strip()
    if not sku:
        return {"status": "error", "message": "SKU必填"}
    base = {
        "views": safe_int(data.get("views")),
        "product_clicks": safe_int(data.get("product_clicks")),
        "orders": safe_int(data.get("orders")),
        "gmv": safe_float(data.get("gmv")),
        "comments": data.get("comments", "").strip()
    }
    diag = diagnose_tiktok_video(base)
    diagnosis = data.get("diagnosis", "").strip() or diag["diagnosis"]
    repeat_action = data.get("repeat_action", "").strip() or diag["repeat_action"]
    conn = get_db()
    try:
        values = (
            data.get("account_name", "").strip(),
            sku,
            data.get("publish_date", "").strip() or datetime.now().strftime("%Y-%m-%d"),
            data.get("video_angle", "").strip(),
            data.get("hook", "").strip(),
            data.get("selling_points", "").strip(),
            data.get("display_order", "").strip(),
            data.get("voiceover", "").strip(),
            data.get("cover_text", "").strip(),
            data.get("caption", "").strip(),
            data.get("hashtags", "").strip(),
            1 if data.get("posted") else 0,
            base["views"],
            base["product_clicks"],
            base["orders"],
            base["gmv"],
            base["comments"],
            diagnosis,
            repeat_action
        )
        if data.get("id"):
            conn.execute("""
                UPDATE tiktok_videos SET
                    account_name=?, sku=?, publish_date=?, video_angle=?, hook=?,
                    selling_points=?, display_order=?, voiceover=?, cover_text=?,
                    caption=?, hashtags=?, posted=?, views=?, product_clicks=?,
                    orders=?, gmv=?, comments=?, diagnosis=?, repeat_action=?,
                    updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            """, values + (safe_int(data.get("id")),))
        else:
            conn.execute("""
                INSERT INTO tiktok_videos
                    (account_name, sku, publish_date, video_angle, hook,
                     selling_points, display_order, voiceover, cover_text,
                     caption, hashtags, posted, views, product_clicks,
                     orders, gmv, comments, diagnosis, repeat_action)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, values)
        conn.commit()
        return {"status": "ok", "diagnosis": diagnosis, "repeat_action": repeat_action}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/api/tiktok/video/delete")
async def tiktok_video_delete(request: Request):
    data = await request.json()
    rid = safe_int(data.get("id"))
    conn = get_db()
    try:
        conn.execute("DELETE FROM tiktok_videos WHERE id=?", (rid,))
        conn.commit()
        return {"status": "ok"}
    finally:
        conn.close()


@app.post("/api/tiktok/script/generate")
async def tiktok_script_generate(request: Request):
    data = await request.json()
    sku = data.get("sku", "").strip()
    angle = data.get("video_angle", "整体效果").strip() or "整体效果"
    conn = get_db()
    try:
        info = find_tiktok_sku_info(conn, sku) if sku else None
        if not info:
            info = {
                "sku": sku,
                "product_name": data.get("product_name", "").strip(),
                "color": data.get("color", "").strip(),
                "length": data.get("length", "").strip(),
                "price": safe_float(data.get("price"), 23.99)
            }
        return {"status": "ok", "sku_info": info, "script": build_tiktok_script(info, angle)}
    finally:
        conn.close()


@app.post("/api/tiktok/schedule/generate")
async def tiktok_schedule_generate(request: Request):
    data = await request.json()
    accounts_raw = data.get("accounts", "")
    if isinstance(accounts_raw, str):
        accounts = [a.strip() for a in accounts_raw.replace("\n", ",").split(",") if a.strip()]
    else:
        accounts = accounts_raw
    start_date = data.get("start_date", "").strip() or datetime.now().strftime("%Y-%m-%d")
    days = safe_int(data.get("days"), 7)
    max_per_sku_per_day = safe_int(data.get("max_per_sku_per_day"), 3)
    selected_skus = data.get("skus") or []

    conn = get_db()
    try:
        all_options = [o for o in get_tiktok_sku_options(conn) if o.get("status") != "paused"]
        if selected_skus:
            selected = {str(s) for s in selected_skus}
            sku_rows = [
                o for o in all_options
                if str(o.get("sku")) in selected or str(o.get("simple_sku")) in selected or str(o.get("tiktok_product_id")) in selected
            ]
        else:
            sku_rows = all_options
        sku_rows = sorted(sku_rows, key=lambda o: (o.get("status") != "main", str(o.get("sku") or "")))[:20]
        rows = generate_tiktok_schedule(sku_rows, accounts, start_date, days, max_per_sku_per_day)
        if data.get("save"):
            for row in rows:
                conn.execute("""
                    INSERT INTO tiktok_videos
                        (account_name, sku, publish_date, video_angle, hook,
                         selling_points, display_order, voiceover, cover_text,
                         caption, hashtags, posted)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0)
                """, (
                    row["account_name"], row["sku"], row["publish_date"], row["video_angle"],
                    row["hook"], row["selling_points"], row["display_order"], row["voiceover"],
                    row["cover_text"], row["caption"], row["hashtags"]
                ))
            conn.commit()
        return {"status": "ok", "rows": rows, "saved": len(rows) if data.get("save") else 0}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


# ────────────────────────────────────────
# 供应商管理
# ────────────────────────────────────────
@app.get("/suppliers", response_class=HTMLResponse)
async def suppliers_page(request: Request):
    """供应商管理页面"""
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM suppliers ORDER BY name").fetchall()
        return render_html("suppliers.html", request, suppliers=[dict(r) for r in rows])
    finally:
        conn.close()


@app.get("/api/suppliers/list")
async def suppliers_list():
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM suppliers ORDER BY name").fetchall()
        return {"suppliers": [dict(r) for r in rows]}
    finally:
        conn.close()


@app.post("/api/suppliers/add")
async def suppliers_add(request: Request):
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        return {"status": "error", "message": "供应商名称必填"}
    conn = get_db()
    try:
        conn.execute("INSERT INTO suppliers (name, seller_id, purchase_link, contact_name, wechat) VALUES (?, ?, ?, ?, ?)",
                     (name, data.get("seller_id", "").strip(), data.get("purchase_link", "").strip(),
                      data.get("contact_name", "").strip(), data.get("wechat", "").strip()))
        conn.commit()
        return {"status": "ok", "message": "已添加"}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/api/suppliers/update")
async def suppliers_update(request: Request):
    data = await request.json()
    rid = data.get("id")
    if not rid:
        return {"status": "error", "message": "缺少ID"}
    conn = get_db()
    try:
        conn.execute("UPDATE suppliers SET name=?, seller_id=?, purchase_link=?, contact_name=?, wechat=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                     (data.get("name", "").strip(), data.get("seller_id", "").strip(),
                      data.get("purchase_link", "").strip(), data.get("contact_name", "").strip(),
                      data.get("wechat", "").strip(), rid))
        conn.commit()
        return {"status": "ok", "message": "已更新"}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/api/suppliers/delete")
async def suppliers_delete(request: Request):
    data = await request.json()
    conn = get_db()
    try:
        conn.execute("DELETE FROM suppliers WHERE id=?", (data.get("id"),))
        conn.commit()
        return {"status": "ok", "message": "已删除"}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


# ────────────────────────────────────────
# 华威尔(宇腾TMS) 连接器
# ────────────────────────────────────────
HUAWELL_BASE = "http://47.115.60.246:8000"
HUAWELL_AUTH = {
    "FACTNO": "003",
    "SUPNO": "HW03359",
    "SUPPASS": "42a36b",
    "APPKEY": "5b62d70582ae2e1b26edb50dbf7d8b36"
}

async def huawell_call(method: str, extra_data: dict = None) -> dict:
    """调用华威尔 API"""
    async with httpx.AsyncClient() as client:
        body = dict(HUAWELL_AUTH)
        if extra_data:
            body.update(extra_data)
        resp = await client.post(
            f"{HUAWELL_BASE}/api/v1/common/{method}",
            json=body,
            timeout=20
        )
        return resp.json()


@app.post("/api/headhaul/sync")
async def sync_headhaul():
    """从华威尔查询头程运单信息"""
    conn = get_db()
    try:
        # 获取已有头程单的运单号,查询最新状态
        existing = conn.execute(
            "SELECT customer_order_no FROM headhaul_orders WHERE customer_order_no != ''"
        ).fetchall()

        synced = 0
        for row in existing:
            packno = row["customer_order_no"]
            if not packno:
                continue
            try:
                result = await huawell_call("getorder", {"PACKNO": packno})
                if result.get("code") == 1:
                    data = result.get("data", [])
                    if isinstance(data, list):
                        data = data[0] if data else {}
                    conn.execute("""
                        UPDATE headhaul_orders SET
                            transfer_order_no = ?, tracking_number = ?,
                            shipping_channel = ?, status = ?,
                            origin = ?, destination = ?,
                            weight = ?, pieces = ?,
                            raw_data = ?, updated_at = CURRENT_TIMESTAMP
                        WHERE customer_order_no = ?
                    """, (
                        str(data.get("zycode", "")),
                        str(data.get("packno", "")),
                        str(data.get("channelname", "")),
                        str(data.get("kdzt", data.get("status", ""))),
                        str(data.get("pol", "")),
                        str(data.get("destnm", "")),
                        float(data.get("yczweit", 0) or 0),
                        int(data.get("jcnt", 1) or 1),
                        json.dumps(data, ensure_ascii=False),
                        packno
                    ))
                    synced += 1
            except Exception as e:
                print(f"[Huawell] Error syncing {packno}: {e}")

        conn.commit()
        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("huawell", "sync_headhaul", "success", f"Updated {synced} records")
        )
        conn.commit()
        return {"status": "ok", "synced": synced}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/api/headhaul/lookup")
async def headhaul_lookup(request: Request):
    """查询单个运单在华威尔的状态"""
    data = await request.json()
    packno = data.get("packno", "").strip()
    if not packno:
        return {"status": "error", "message": "请输入运单号"}

    conn = get_db()
    try:
        result = await huawell_call("getorder", {"PACKNO": packno})
        if result.get("code") != 1:
            return {"status": "error", "message": result.get("msg", "查询失败")}

        order_data = result.get("data", [])
        if isinstance(order_data, list):
            order_data = order_data[0] if order_data else {}

        # 存入数据库
        conn.execute("""
            INSERT OR REPLACE INTO headhaul_orders
            (customer_order_no, transfer_order_no, tracking_number, shipping_channel,
             status, origin, destination, weight, pieces, raw_data, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        """, (
            str(order_data.get("supcode", order_data.get("packno", packno))),
            str(order_data.get("zycode", "")),
            str(order_data.get("packno", packno)),
            str(order_data.get("channelname", "")),
            str(order_data.get("kdzt", order_data.get("status", "未知"))),
            str(order_data.get("pol", "")),
            str(order_data.get("destnm", "")),
            float(order_data.get("yczweit", 0) or 0),
            int(order_data.get("jcnt", 1) or 1),
            json.dumps(order_data, ensure_ascii=False)
        ))
        conn.commit()

        return {"status": "ok", "data": order_data}
    finally:
        conn.close()


@app.post("/api/headhaul/tracking")
async def headhaul_tracking(request: Request):
    """查询运单轨迹"""
    data = await request.json()
    packno = data.get("packno", "").strip()
    if not packno:
        return {"status": "error", "message": "请输入运单号"}

    result = await huawell_call("tracking", {"PACKNO": [packno]})
    return result


@app.post("/api/headhaul/channels")
async def headhaul_channels():
    """获取可用渠道类型"""
    result = await huawell_call("getbasedata", {"TYPENO": "QUDAO"})
    return result


@app.post("/api/headhaul/create")
async def headhaul_create(request: Request):
    """创建头程运单"""
    data = await request.json()
    # 构建订单创建请求
    order_data = {
        "data": {
            "Factno": HUAWELL_AUTH["FACTNO"],
            "supno": HUAWELL_AUTH["SUPNO"],
            "suppass": HUAWELL_AUTH["SUPPASS"],
            "appkey": HUAWELL_AUTH["APPKEY"],
            "channeltype": data.get("channeltype", ""),
            "dest": data.get("dest", "US"),
            "expresstype": data.get("expresstype", "04"),
            "delcompanynm": data.get("delcompanynm", ""),
            "contactsby": data.get("contactsby", ""),
            "address1": data.get("address1", ""),
            "phone": data.get("phone", ""),
            "postno": data.get("postno", ""),
            "state": data.get("state", ""),
            "city": data.get("city", ""),
            "smatno": data.get("smatno", ""),
            "bxflg": data.get("bxflg", "N"),
            "declaration": data.get("declaration", "01"),
            "clearance": data.get("clearance", ""),
            "packno": data.get("packno", ""),
            "supcode": data.get("supcode", data.get("packno", "")),
            "organo": data.get("organo", ""),
            "worgno": data.get("worgno", ""),
            "productname": data.get("productname", ""),
            "addnm": data.get("addnm", ""),
            "address2": data.get("address2", ""),
            "address3": data.get("address3", ""),
            "mailNo": data.get("mailNo", ""),
            "indamt": data.get("indamt", "0.00"),
            "remark": data.get("remark", ""),
            "yfdfMk": data.get("yfdfMk", ""),
            "dfdw": data.get("dfdw", ""),
            "dfCash": data.get("dfCash", 0),
            "dcis_type": data.get("dcis_type", ""),
            "upsMk": data.get("upsMk", ""),
            "sendcountry": data.get("sendcountry", ""),
            "sdcompanynm": data.get("sdcompanynm", ""),
            "sendby": data.get("sendby", ""),
            "sendbyaddress": data.get("sendbyaddress", ""),
            "sendmobi": data.get("sendmobi", ""),
            "sendpostno": data.get("sendpostno", ""),
            "sedprovinces": data.get("sedprovinces", ""),
            "sedcity": data.get("sedcity", ""),
            "sendadd2": data.get("sendadd2", ""),
            "sendadd3": data.get("sendadd3", ""),
            "sendaddnm": data.get("sendaddnm", ""),
            "vatno": data.get("vatno", ""),
            "eorino": data.get("eorino", ""),
            "vatfact": data.get("vatfact", ""),
            "vatadd": data.get("vatadd", ""),
            "vatphone": data.get("vatphone", ""),
            "kacheNo": data.get("kacheNo", ""),
            "fbano": data.get("fbano", ""),
            "shipmentID": data.get("shipmentID", ""),
            "deliverymode": data.get("deliverymode", ""),
            "zycode": data.get("zycode", ""),
            "fileurl1": data.get("fileurl1", ""),
            "fileurl2": data.get("fileurl2", ""),
            "fileurl3": data.get("fileurl3", ""),
            "typeno": data.get("typeno", ""),
            "channelprice": data.get("channelprice", ""),
            "orderDetailList": data.get("orderDetailList", [{
                "zhuangtai": "OT",
                "mlength": data.get("mlength", 1),
                "mwidth": data.get("mwidth", 1),
                "mheight": data.get("mheight", 1),
                "mweit": data.get("mweit", "1.00"),
                "jcnt": data.get("jcnt", 1),
                "jobno": data.get("jobno", data.get("packno", "1")),
                "refno": data.get("refno", ""),
                "ordersmatlist": data.get("ordersmatlist", [{
                    "jobid": data.get("jobno", data.get("packno", "1")),
                    "qty": data.get("qty", 1),
                    "value": data.get("value", "1.00"),
                    "productch": data.get("productch", ""),
                    "producten": data.get("producten", ""),
                    "pinpai": data.get("pinpai", ""),
                    "pptype": data.get("pptype", ""),
                    "hscode": data.get("hscode", ""),
                    "xinghao": data.get("xinghao", ""),
                    "caizhi": data.get("caizhi", ""),
                    "yongtu": data.get("yongtu", ""),
                    "baozhuang": data.get("baozhuang", ""),
                    "jingweit": data.get("jingweit", "0.00"),
                    "daino": data.get("daino", ""),
                    "skuno": data.get("skuno", ""),
                    "caigouprice": data.get("caigouprice", "0.00"),
                    "forcustno": data.get("forcustno", ""),
                    "bgbibie": data.get("bgbibie", ""),
                    "qgbibie": data.get("qgbibie", "USD"),
                    "bgvalue": data.get("bgvalue", "0.00"),
                    "salesvalue": data.get("salesvalue", "0.00"),
                    "zwmat": data.get("zwmat", ""),
                    "zwuse": data.get("zwuse", ""),
                    "zwarticno": data.get("zwarticno", ""),
                    "salesurl": data.get("salesurl", ""),
                    "saveurl": data.get("saveurl", ""),
                    "picbase64": data.get("picbase64", ""),
                    "smatnos": data.get("smatnos", "")
                }])
            }])
        }
    }

    async with httpx.AsyncClient() as client:
        resp = await client.post(
            f"{HUAWELL_BASE}/api/v1/common/createorder",
            json=order_data,
            timeout=30
        )
        result = resp.json()

        if result.get("code") == 1:
            # 存入数据库
            conn = get_db()
            try:
                d = result.get("data", [{}])[0] if result.get("data") else {}
                conn.execute("""
                    INSERT OR REPLACE INTO headhaul_orders
                    (customer_order_no, transfer_order_no, tracking_number, shipping_channel,
                     status, origin, destination, pieces, raw_data, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                """, (
                    str(d.get("packno", data.get("packno", ""))),
                    str(d.get("trackingNo", "")),
                    str(d.get("packno", "")),
                    str(data.get("channeltype", "")),
                    "已创建",
                    str(data.get("sendcountry", "")),
                    str(data.get("dest", "")),
                    int(d.get("jcnt", 1)),
                    json.dumps(result, ensure_ascii=False)
                ))
                conn.commit()
            finally:
                conn.close()

        return result


# ────────────────────────────────────────
# 物流追踪
# ────────────────────────────────────────
@app.post("/api/tracking/add")
async def add_tracking(request: Request):
    """手动添加物流单号"""
    data = await request.json()
    order_id = data.get("order_id")
    tracking_number = data.get("tracking_number")
    tracking_company = data.get("tracking_company", "云途物流")

    if not order_id or not tracking_number:
        return {"status": "error", "message": "缺少 order_id 或 tracking_number"}

    conn = get_db()
    try:
        conn.execute("""
            UPDATE orders SET tracking_number = ?, tracking_company = ?,
            status = 'shipped', updated_at = CURRENT_TIMESTAMP
            WHERE platform_order_id = ? OR id = ?
        """, (tracking_number, tracking_company, str(order_id), order_id if str(order_id).isdigit() else 0))

        if conn.total_changes == 0:
            return {"status": "error", "message": "未找到该订单"}

        conn.commit()
        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("manual", "add_tracking", "success",
             f"Order #{order_id} tracking: {tracking_number} via {tracking_company}")
        )
        conn.commit()
        return {"status": "ok", "message": "运单号已录入"}
    finally:
        conn.close()


@app.get("/api/tracking/lookup/{tracking_number}")
async def lookup_tracking(tracking_number: str):
    """查询物流轨迹（通过 17TRACK）"""
    import httpx
    conn = get_db()
    try:
        order = conn.execute(
            "SELECT * FROM orders WHERE tracking_number = ?",
            (tracking_number,)
        ).fetchone()

        if not order:
            return {"status": "error", "message": "未找到该运单号对应的订单"}

        # 尝试通过 17TRACK API 查询
        try:
            async with httpx.AsyncClient() as client:
                resp = await client.get(
                    f"https://api.17track.net/track/v2.2/gettrackinfo",
                    params={"numbers": tracking_number},
                    headers={"17token": os.getenv("TRACK17_TOKEN", "")},
                    timeout=10
                )
                if resp.status_code == 200:
                    track_data = resp.json()
                    return {
                        "status": "ok",
                        "order": dict(order),
                        "tracking": track_data
                    }
        except Exception:
            pass

        # 回退：返回基本信息 + 17TRACK 链接
        return {
            "status": "ok",
            "order": dict(order),
            "tracking_number": tracking_number,
            "tracking_company": order["tracking_company"],
            "track_url": f"https://t.17track.net/zh-cn#nums={tracking_number}"
        }
    finally:
        conn.close()


@app.post("/api/tracking/batch")
async def batch_add_tracking(request: Request):
    """批量录入运单号 [{\"order_id\": \"xxx\", \"tracking_number\": \"YT001\"}, ...]"""
    data = await request.json()
    if not isinstance(data, list):
        return {"status": "error", "message": "需要数组格式"}

    conn = get_db()
    success = 0
    failed = 0
    try:
        for item in data:
            oid = item.get("order_id")
            tn = item.get("tracking_number")
            tc = item.get("tracking_company", "云途物流")
            if oid and tn:
                conn.execute("""
                    UPDATE orders SET tracking_number = ?, tracking_company = ?,
                    status = 'shipped', updated_at = CURRENT_TIMESTAMP
                    WHERE platform_order_id = ? OR id = ?
                """, (tn, tc, str(oid), oid if str(oid).isdigit() else 0))
                if conn.total_changes > 0:
                    success += 1
                else:
                    failed += 1
        conn.commit()
        return {"status": "ok", "success": success, "failed": failed}
    finally:
        conn.close()


# ────────────────────────────────────────
# 重要事项提醒
# ────────────────────────────────────────
@app.get("/matters", response_class=HTMLResponse)
async def matters_page(request: Request):
    conn = get_db()
    try:
        matters = conn.execute(
            "SELECT * FROM important_matters ORDER BY status ASC, id DESC"
        ).fetchall()
        return render_html("matters.html", request, matters=[dict(m) for m in matters])
    finally:
        conn.close()


@app.post("/api/matters/add")
async def add_matter(request: Request):
    data = await request.json()
    title = data.get("title", "").strip()
    content = data.get("content", "").strip()
    remind_at = data.get("remind_at", "").strip()
    if not title or not remind_at:
        return {"status": "error", "message": "标题和提醒时间必填"}
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO important_matters (title, content, remind_at) VALUES (?, ?, ?)",
            (title, content, remind_at)
        )
        conn.commit()
        return {"status": "ok"}
    finally:
        conn.close()


@app.post("/api/matters/complete")
async def complete_matter(request: Request):
    data = await request.json()
    mid = data.get("id")
    conn = get_db()
    try:
        conn.execute(
            "UPDATE important_matters SET status='done', updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (mid,)
        )
        conn.commit()
        return {"status": "ok"}
    finally:
        conn.close()


@app.post("/api/matters/delete")
async def delete_matter(request: Request):
    data = await request.json()
    mid = data.get("id")
    conn = get_db()
    try:
        conn.execute("DELETE FROM important_matters WHERE id=?", (mid,))
        conn.commit()
        return {"status": "ok"}
    finally:
        conn.close()


@app.get("/api/matters/check")
async def check_matters():
    """检查是否有到期的提醒"""
    conn = get_db()
    try:
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        due = conn.execute(
            "SELECT * FROM important_matters WHERE status='pending' AND remind_at <= ? ORDER BY remind_at ASC",
            (now,)
        ).fetchall()
        return {"due": [dict(m) for m in due]}
    finally:
        conn.close()


# ────────────────────────────────────────
# 邮件提醒
# ────────────────────────────────────────

@app.get("/reminders", response_class=HTMLResponse)
async def reminders_page(request: Request):
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM reminders ORDER BY remind_date ASC, id DESC").fetchall()
        items = [dict(r) for r in rows]
        return render_html("reminders.html", request, items=items, today=datetime.now().strftime("%Y-%m-%d"))
    finally:
        conn.close()


@app.post("/api/reminders/add")
async def reminders_add(request: Request):
    data = await request.json()
    title = data.get("title", "").strip()
    content = data.get("content", "").strip()
    remind_date = data.get("remind_date", "").strip()
    email = data.get("email", "").strip()
    repeat_type = data.get("repeat_type", "").strip()
    repeat_day = int(data.get("repeat_day", 0) or 0)
    if not title or not remind_date:
        return {"status": "error", "message": "标题和日期必填"}
    # 从日期自动推导repeat_day
    if repeat_type and not repeat_day:
        from datetime import datetime as dt
        d = dt.strptime(remind_date, "%Y-%m-%d")
        if repeat_type == "monthly":
            repeat_day = d.day  # 几号
        elif repeat_type == "weekly":
            repeat_day = d.weekday()  # 0=周一
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO reminders (title, content, remind_date, email, repeat_type, repeat_day)"
            " VALUES (?, ?, ?, ?, ?, ?)",
            (title, content, remind_date, email or EMAIL_TO, repeat_type, repeat_day)
        )
        conn.commit()
        return {"status": "ok", "id": conn.execute("SELECT last_insert_rowid()").fetchone()[0]}
    finally:
        conn.close()


@app.post("/api/reminders/delete")
async def reminders_delete(request: Request):
    data = await request.json()
    rid = data.get("id")
    conn = get_db()
    try:
        conn.execute("DELETE FROM reminders WHERE id=?", (rid,))
        conn.commit()
        return {"status": "ok"}
    finally:
        conn.close()


@app.post("/api/reminders/test-email")
async def reminders_test_email():
    """测试邮件发送"""
    ok = send_email(
        to=EMAIL_TO,
        subject="🧪 LuxeLocks Hub 邮件测试",
        body="如果你收到这封邮件，说明邮件提醒功能配置成功！"
    )
    return {"status": "ok" if ok else "error", "message": "测试邮件已发送" if ok else "发送失败，请检查SMTP配置"}


# ────────────────────────────────────────
# 产品媒体文件管理
# ────────────────────────────────────────

@app.get("/media", response_class=HTMLResponse)
async def media_page(request: Request):
    resp = render_html("media.html", request, media_root=str(MEDIA_ROOT))
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


@app.get("/api/media/tree")
async def media_tree(force: bool = False):
    """返回所有文件/文件夹路径（5分钟缓存）"""
    import subprocess, time
    from concurrent.futures import ThreadPoolExecutor
    global _media_tree_cache
    # 缓存命中且未过期
    if not force and _media_tree_cache["paths"] and (time.time() - _media_tree_cache["ts"]) < _MEDIA_CACHE_TTL:
        return {"paths": _media_tree_cache["paths"], "cached": True}
    paths = []
    try:
        import tempfile, os as _os
        root = str(_get_media_root())
        root_len = len(root)
        def list_items(what):
            # PowerShell写出UTF-8文件，Python直接读取（彻底避免编码问题）
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".txt")
            tmp.close()
            try:
                script = f"Get-ChildItem -Path '{root}' -Recurse -{what} -ErrorAction SilentlyContinue | ForEach-Object {{ $_.FullName }} | Out-File -FilePath '{tmp.name}' -Encoding utf8"
                subprocess.run(
                    ["powershell", "-NoProfile", "-Command", script],
                    capture_output=True, timeout=60
                )
                with open(tmp.name, "r", encoding="utf-8") as f:
                    text = f.read()
            finally:
                _os.unlink(tmp.name)
            items = []
            for line in text.strip().split("\n"):
                line = line.strip()
                if line and len(line) > root_len:
                    rel = line[root_len:].lstrip("\\").lstrip("/")
                    items.append({"name": rel.replace("\\", "/"), "is_dir": what == "Directory"})
            return items
        with ThreadPoolExecutor(max_workers=2) as pool:
            f_files = pool.submit(list_items, "File")
            f_dirs = pool.submit(list_items, "Directory")
            paths = f_files.result() + f_dirs.result()
    except Exception:
        pass
    # 保存到缓存
    _media_tree_cache = {"paths": paths, "ts": time.time()}
    return {"paths": paths, "cached": False}


@app.get("/api/media/by-sku")
async def media_by_sku(sku: str = ""):
    """按简化SKU查找本地素材图片（使用缓存树，毫秒级响应）"""
    if not sku:
        return {"status": "error", "message": "缺少sku参数"}
    global _media_tree_cache
    # 如果缓存为空，先加载
    if not _media_tree_cache["paths"]:
        await media_tree()
    images = []
    sku_lower = sku.lower()
    image_exts = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.mp4', '.mov', '.avi'}
    for item in _media_tree_cache["paths"]:
        if item["is_dir"]:
            continue
        if sku_lower in item["name"].lower():
            ext = Path(item["name"]).suffix.lower()
            if ext in image_exts:
                images.append(item["name"])
    return {"status": "ok", "sku": sku, "images": images}


def _safe_path(rel: str) -> Path:
    """防止路径穿越攻击"""
    clean = rel.replace("\\", "/").strip("/")
    p = (MEDIA_ROOT / clean).resolve() if clean else _get_media_root()
    if not str(p).startswith(str(MEDIA_ROOT)):
        raise HTTPException(403, "Access denied")
    return p


@app.get("/api/media/list")
async def media_list(path: str = ""):
    target = _safe_path(path)
    if not target.exists() or not target.is_dir():
        target = _get_media_root()
        path = ""
    items = []
    for entry in sorted(target.iterdir(), key=lambda e: (not e.is_dir(), e.name.lower())):
        stat = entry.stat()
        items.append({
            "name": entry.name,
            "is_dir": entry.is_dir(),
            "size": stat.st_size if entry.is_file() else 0,
            "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
            "ext": entry.suffix.lower() if entry.is_file() else "",
        })
    # breadcrumb
    parts = path.replace("\\", "/").strip("/").split("/") if path else []
    breadcrumb = [{"name": "根目录", "path": ""}]
    acc = ""
    for p in parts:
        if p:
            acc = (acc + "/" + p).lstrip("/")
            breadcrumb.append({"name": p, "path": acc})
    return {"items": items, "breadcrumb": breadcrumb, "current": path}


@app.post("/api/media/mkdir")
async def media_mkdir(request: Request):
    data = await request.json()
    path = data.get("path", "")
    name = data.get("name", "").strip()
    if not name:
        return {"status": "error", "message": "文件夹名必填"}
    target = _safe_path(path) / name
    if target.exists():
        return {"status": "error", "message": "已存在"}
    target.mkdir(parents=True)
    return {"status": "ok"}


@app.post("/api/media/delete")
async def media_delete(request: Request):
    data = await request.json()
    path = data.get("path", "")
    names = data.get("names", [])  # 支持批量删除
    if not names:
        return {"status": "error", "message": "请选择文件"}
    deleted = 0
    for name in names:
        target = _safe_path(path) / name
        if not target.exists():
            continue
        if target.is_dir():
            shutil.rmtree(target)
        else:
            target.unlink()
        deleted += 1
    return {"status": "ok", "deleted": deleted}


@app.post("/api/media/upload")
async def media_upload(path: str = "", files: list[UploadFile] = File(...)):
    target_dir = _safe_path(path)
    target_dir.mkdir(parents=True, exist_ok=True)
    uploaded = []
    for file in files:
        safe_name = Path(file.filename).name  # 防路径穿越
        dest = target_dir / safe_name
        content = await file.read()
        dest.write_bytes(content)
        uploaded.append(safe_name)
    return {"status": "ok", "uploaded": uploaded}


@app.get("/api/media/preview")
async def media_preview(path: str = ""):
    """返回图片文件用于预览"""
    target = _safe_path(path)
    if not target.exists() or not target.is_file():
        raise HTTPException(404)
    from fastapi.responses import FileResponse
    return FileResponse(target)


@app.get("/api/img/proxy")
async def img_proxy(url: str = ""):
    """代理远程图片，本地磁盘缓存（首次curl拉取，后续直接读缓存）"""
    if not url or not url.startswith("http"):
        raise HTTPException(400)
    import subprocess, hashlib, os
    from fastapi.responses import FileResponse, Response
    # 本地缓存：URL的MD5作为文件名
    cache_key = hashlib.md5(url.encode()).hexdigest()
    cache_path = _IMG_CACHE_DIR / cache_key
    # 命中缓存
    if cache_path.exists() and cache_path.stat().st_size > 100:
        return FileResponse(cache_path, media_type="image/jpeg",
            headers={"Cache-Control": "public, max-age=86400"})
    # 拉取并缓存
    try:
        r = subprocess.run(
            ["curl", "-s", "-L", "--connect-timeout", "10", "--max-time", "20", "-o", str(cache_path), url],
            capture_output=True, timeout=25
        )
        if r.returncode == 0 and cache_path.stat().st_size > 100:
            return FileResponse(cache_path, media_type="image/jpeg",
                headers={"Cache-Control": "public, max-age=86400"})
        else:
            cache_path.unlink(missing_ok=True)
            raise HTTPException(502, detail=f"curl failed, code={r.returncode}")
    except HTTPException:
        raise
    except Exception as e:
        cache_path.unlink(missing_ok=True)
        raise HTTPException(502, detail=str(e)[:200])


# ────────────────────────────────────────
# 库存管理
# ────────────────────────────────────────

@app.get("/inventory/new-products", response_class=HTMLResponse)
async def inventory_new_products(request: Request):
    """新品页 - 手动录入新品SKU"""
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM warehouse_inventory WHERE warehouse_name='NewProducts' ORDER BY reference_code"
        ).fetchall()
        items = [dict(i) for i in rows]
        totals = {
            "available": sum(i["available_inventory"] or 0 for i in items),
            "transit": sum(i["in_transit_total"] or 0 for i in items),
        }
        return render_html("inventory_new_products.html", request,
            items=items, totals=totals)
    finally:
        conn.close()


@app.post("/api/inventory/new-products/add")
async def new_products_add(request: Request):
    data = await request.json()
    sku = data.get("sku", "").strip()
    available = int(data.get("available", 0) or 0)
    transit = int(data.get("transit", 0) or 0)
    if not sku:
        return {"status": "error", "message": "SKU必填"}
    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO warehouse_inventory (reference_code, warehouse_name, available_inventory, in_transit_total)
            VALUES (?, 'NewProducts', ?, ?)
        """, (sku, available, transit))
        conn.commit()
        return {"status": "ok", "id": conn.execute("SELECT last_insert_rowid()").fetchone()[0]}
    finally:
        conn.close()


@app.post("/api/inventory/new-products/delete")
async def new_products_delete(request: Request):
    data = await request.json()
    rid = data.get("id")
    conn = get_db()
    try:
        conn.execute("DELETE FROM warehouse_inventory WHERE id=? AND warehouse_name='NewProducts'", (rid,))
        conn.commit()
        return {"status": "ok"}
    finally:
        conn.close()


@app.post("/api/inventory/new-products/clear")
async def new_products_clear():
    conn = get_db()
    try:
        conn.execute("DELETE FROM warehouse_inventory WHERE warehouse_name='NewProducts'")
        conn.commit()
        return {"status": "ok"}
    finally:
        conn.close()


@app.get("/inventory/luxelocks", response_class=HTMLResponse)
async def inventory_luxelocks(request: Request, sort: str = None, order: str = "desc"):
    """LuxeLocks仓库存"""
    conn = get_db()
    try:
        sort_col = "available_inventory" if sort == "available" else ("in_transit_total" if sort == "transit" else "reference_code")
        sort_dir = "ASC" if order == "asc" else "DESC"
        rows = conn.execute(
            f"SELECT * FROM warehouse_inventory WHERE warehouse_name = ? ORDER BY {sort_col} {sort_dir}",
            ("LuxeLocks",)
        ).fetchall()
        items = [dict(i) for i in rows]
        totals = {
            "available": sum(i["available_inventory"] or 0 for i in items),
            "transit": sum(i["in_transit_total"] or 0 for i in items),
        }
        return render_html("inventory_warehouse.html", request,
            items=items, totals=totals, sort=sort, order=order,
            warehouse_name="LuxeLocks仓",
            warehouse_key="luxelocks")
    finally:
        conn.close()


# ────────────────────────────────────────
# SKU 映射管理
# ────────────────────────────────────────
@app.get("/sku-mapping", response_class=HTMLResponse)
async def sku_mapping_page(request: Request):
    conn = get_db()
    try:
        mappings = conn.execute("SELECT * FROM sku_mapping ORDER BY shopify_simple_sku").fetchall()
        return render_html("sku_mapping.html", request, mappings=[dict(m) for m in mappings])
    finally:
        conn.close()


@app.post("/api/sku-mapping/save")
async def sku_mapping_save(request: Request):
    data = await request.json()
    shopify = data.get("shopify_simple_sku", "").strip()
    cross = data.get("cross_border_simple_sku", "").strip()
    luxe = data.get("luxelocks_simple_sku", "").strip()
    veloura = data.get("velourahair_simple_sku", "").strip()

    if not shopify:
        return {"status": "error", "message": "Shopify简化SKU必填"}

    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO sku_mapping (shopify_simple_sku, cross_border_simple_sku, luxelocks_simple_sku, velourahair_simple_sku, updated_at)
            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(shopify_simple_sku) DO UPDATE SET
                cross_border_simple_sku = excluded.cross_border_simple_sku,
                luxelocks_simple_sku = excluded.luxelocks_simple_sku,
                velourahair_simple_sku = excluded.velourahair_simple_sku,
                updated_at = CURRENT_TIMESTAMP
        """, (shopify, cross or None, luxe or None, veloura or None))
        conn.commit()
        return {"status": "ok", "message": "已保存"}
    finally:
        conn.close()


@app.post("/api/sku-mapping/delete")
async def sku_mapping_delete(request: Request):
    data = await request.json()
    mid = data.get("id")
    conn = get_db()
    try:
        conn.execute("DELETE FROM sku_mapping WHERE id = ?", (mid,))
        conn.commit()
        return {"status": "ok"}
    finally:
        conn.close()


@app.get("/inventory/velourahair", response_class=HTMLResponse)
async def inventory_velourahair(request: Request, sort: str = None, order: str = "desc"):
    """VelouraHair仓库存"""
    conn = get_db()
    try:
        sort_col = "available_inventory" if sort == "available" else ("in_transit_total" if sort == "transit" else "reference_code")
        sort_dir = "ASC" if order == "asc" else "DESC"
        rows = conn.execute(
            f"SELECT * FROM warehouse_inventory WHERE warehouse_name = ? ORDER BY {sort_col} {sort_dir}",
            ("VelouraHair",)
        ).fetchall()
        items = [dict(i) for i in rows]
        totals = {
            "available": sum(i["available_inventory"] or 0 for i in items),
            "transit": sum(i["in_transit_total"] or 0 for i in items),
        }
        return render_html("inventory_warehouse.html", request,
            items=items, totals=totals, sort=sort, order=order,
            warehouse_name="VelouraHair仓",
            warehouse_key="velourahair")
    finally:
        conn.close()


@app.post("/api/inventory/import")
async def import_warehouse_inventory(request: Request):
    """导入xlsx库存表，清空历史并更新"""
    try:
        import openpyxl
    except ImportError:
        return {"status": "error", "message": "需要安装 openpyxl: pip install openpyxl"}

    form = await request.form()
    warehouse = form.get("warehouse", "").strip()
    file = form.get("file")

    if not warehouse:
        return {"status": "error", "message": "缺少warehouse参数"}
    if not file:
        return {"status": "error", "message": "请上传xlsx文件"}

    try:
        # 读取xlsx
        contents = await file.read()
        import io
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active

        # 读取表头行，建立列映射
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx

        # 自动检测xlsx格式: 新格式(SKU) 或 旧格式(Reference code)
        use_new_format = any("SKU" == str(k).strip() for k in headers)
        use_old_format = any("Reference code" in str(k) for k in headers)

        if not use_new_format and not use_old_format:
            return {"status": "error", "message": "缺少 SKU 或 Reference code 列"}

        # 模糊匹配: 在表头中查找包含关键词的列名
        def find_col(*keywords):
            for kw in keywords:
                for h in headers:
                    if kw.lower() in h.lower():
                        return h
            return None

        avail_col = find_col("Available Inventory", "Available inventory", "可用库存")
        transit_col = find_col("In-transit inventory", "In Transit", "在途库存")

        if not avail_col or not transit_col:
            return {"status": "error", "message": f"缺少可用库存或在途库存列. 表头: {list(headers.keys())[:10]}"}

        # 读取所有数据行
        def get_val(row, col_name, default=""):
            idx = headers.get(col_name)
            if idx:
                val = row[idx - 1]
                return str(val).strip() if val is not None else default
            return default

        # 安全转整数
        def safe_int(val, default=0):
            try:
                return int(float(val))
            except (ValueError, TypeError):
                return default

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(cell for cell in row):
                continue

            if use_new_format:
                ref_code = get_val(row, "SKU")
            else:
                ref_code = get_val(row, "Reference code")
            if not ref_code:
                continue

            available = safe_int(get_val(row, avail_col, "0"))
            transit = safe_int(get_val(row, transit_col, "0"))

            # 新格式: 在途 = In-transit inventory (已经是净值)
            # 旧格式: 在途 = In Transit Total - In Transit Returning
            if use_old_format:
                returning = safe_int(get_val(row, "In Transit: Returning Quantity"))
                transit = max(0, transit - returning)

            rows.append({
                "reference_code": ref_code,
                "warehouse_name": warehouse,
                "available_inventory": available,
                "in_transit_total": transit,
                "goods_id": get_val(row, "Goods ID") if use_old_format else "",
                "goods_name": get_val(row, "Goods name") if use_old_format else get_val(row, "Product Name"),
                "status": get_val(row, "Status") if use_old_format else get_val(row, "Stock Property"),
                "total_inventory": safe_int(get_val(row, "Total inventory")) if use_old_format else safe_int(get_val(row, "Total Stock")),
                "defective": safe_int(get_val(row, "Defective")),
                "good_reserved": safe_int(get_val(row, "Good&Reserved")),
                "in_transit_shipped": 0,
                "in_transit_receiving": 0,
                "in_transit_returning": 0,
                "platform": get_val(row, "Platform") if use_old_format else get_val(row, "Warehouse"),
            })
    except Exception as e:
        return {"status": "error", "message": f"解析xlsx失败: {str(e)}"}

    if not rows:
        return {"status": "error", "message": "xlsx中没有有效数据行"}

    conn = get_db()
    try:
        # 按 Reference code 汇总
        aggregated = {}
        for r in rows:
            ref = r["reference_code"]
            if ref not in aggregated:
                aggregated[ref] = {
                    "goods_id": r["goods_id"],
                    "goods_name": r["goods_name"],
                    "reference_code": ref,
                    "status": r["status"],
                    "warehouse_name": warehouse,
                    "total_inventory": 0,
                    "available_inventory": 0,
                    "defective": 0,
                    "good_reserved": 0,
                    "in_transit_total": 0,
                    "in_transit_shipped": 0,
                    "in_transit_receiving": 0,
                    "in_transit_returning": 0,
                    "platform": r["platform"],
                }
            agg = aggregated[ref]
            agg["total_inventory"] += r["total_inventory"]
            agg["available_inventory"] += r["available_inventory"]
            agg["defective"] += r["defective"]
            agg["good_reserved"] += r["good_reserved"]
            agg["in_transit_total"] += r["in_transit_total"]
            agg["in_transit_shipped"] += r["in_transit_shipped"]
            agg["in_transit_receiving"] += r["in_transit_receiving"]
            agg["in_transit_returning"] += r["in_transit_returning"]

        # 净在途 = 在途总量 - 退货在途
        for ref, agg in aggregated.items():
            agg["in_transit_total"] = max(0, agg["in_transit_total"] - agg["in_transit_returning"])

        # 清空该仓库历史数据
        conn.execute("DELETE FROM warehouse_inventory WHERE warehouse_name = ?", (warehouse,))
        # 插入汇总后的数据
        for ref, r in aggregated.items():
            conn.execute("""
                INSERT INTO warehouse_inventory
                (goods_id, goods_name, reference_code, status, warehouse_name,
                 total_inventory, available_inventory, defective, good_reserved,
                 in_transit_total, in_transit_shipped, in_transit_receiving, in_transit_returning, platform)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                r["goods_id"], r["goods_name"], r["reference_code"], r["status"], r["warehouse_name"],
                r["total_inventory"], r["available_inventory"], r["defective"], r["good_reserved"],
                r["in_transit_total"], r["in_transit_shipped"], r["in_transit_receiving"], r["in_transit_returning"],
                r["platform"]
            ))
        conn.commit()
        conn.execute(
            "INSERT INTO sync_log (platform, action, status, detail) VALUES (?, ?, ?, ?)",
            ("warehouse", "import_xlsx", "success", f"{warehouse}: {len(rows)} rows → {len(aggregated)} SKUs")
        )
        conn.commit()
        return {"status": "ok", "imported": len(aggregated), "total_rows": len(rows), "warehouse": warehouse}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        conn.close()


@app.post("/api/inventory/clear")
async def clear_warehouse_inventory(request: Request):
    """清空指定仓库库存"""
    data = await request.json()
    warehouse = data.get("warehouse", "").strip()
    if not warehouse:
        return {"status": "error", "message": "缺少warehouse参数"}
    conn = get_db()
    try:
        conn.execute("DELETE FROM warehouse_inventory WHERE warehouse_name = ?", (warehouse,))
        conn.commit()
        return {"status": "ok", "message": f"已清空 {warehouse} 库存数据"}
    finally:
        conn.close()


# ────────────────────────────────────────
# 健康检查
# ────────────────────────────────────────
@app.get("/health")
async def health():
    return {"status": "ok", "timestamp": datetime.now().isoformat()}


if __name__ == "__main__":
    import uvicorn
    import logging
    logging.basicConfig(
        filename=str(BASE_DIR / "app.log"),
        level=logging.INFO,
        format='%(asctime)s %(levelname)s %(message)s'
    )
    logger = logging.getLogger(__name__)

    while True:
        try:
            uvicorn.run(app, host="0.0.0.0", port=8001, log_level="info")
        except Exception as e:
            logger.error(f"Server crashed: {e}, restarting in 3s...")
            import time
            time.sleep(3)
